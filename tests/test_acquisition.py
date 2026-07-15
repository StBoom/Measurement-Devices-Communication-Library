from __future__ import annotations

import concurrent.futures
import sys
import tempfile
import unittest
from types import SimpleNamespace
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from instrument_visa.acquisition import (  # noqa: E402
    AcquisitionResult,
    capture_screenshot,
    capture_sparameters,
    capture_waveform,
    read_power_supply_settings,
    read_signal_generator_settings,
    read_scope_measurement,
    read_value,
    set_power_supply,
    set_power_supply_master_output,
    set_power_supply_output,
    set_signal_generator,
    set_signal_generator_rf_output,
)
from instrument_visa.config import SParameterConfig  # noqa: E402
from instrument_visa.excel_export import append_result  # noqa: E402
from instrument_visa.cli import _list_all_resources, _load_sequence_config  # noqa: E402
from instrument_visa.picoscope_client import parse_pico_analog_channels, parse_pico_digital_channels, picoscope_analog_channels_for_variant, picoscope_variant_supports_digital  # noqa: E402
from instrument_visa.saleae_client import SaleaeCanConfig, SaleaeCaptureConfig, SaleaeI2cConfig, SaleaeSpiConfig, SaleaeUartConfig, parse_saleae_channels  # noqa: E402
from instrument_visa.profiles import detect_profile  # noqa: E402
from instrument_visa.visa_client import VisaInstrument  # noqa: E402
from instrument_visa.sequence import (  # noqa: E402
    CustomSequenceConfig,
    CA410Config,
    MAX_SWEEP_POINTS,
    FrequencySweepConfig,
    SequenceStep,
    SequenceVariable,
    TimedSwitchConfig,
    VoltageSweepConfig,
    frequency_points,
    parse_json_bool,
    _bool_param,
    parse_frequency_hz,
    parse_ssh_target,
    probe_direct_serial_idn,
    parse_serial_format,
    parse_parallel_tasks,
    read_direct_serial_log,
    send_direct_serial_command,
    parse_34970a_channels,
    read_34970a_data_logger,
    DataLogger34970AConfig,
    parse_ca410_measurement_response,
    read_ca410_measurement,
    run_ssh_command,
    parse_34970a_measurement_plan,
    run_custom_sequence,
    run_frequency_sweep,
    run_timed_switch,
    run_voltage_sweep,
    sanitized_step_params,
    voltage_points,
)


class FakeInstrument:
    def __init__(self, query_responses: dict[str, str] | None = None, binary_responses: dict[str, bytes] | None = None) -> None:
        self.address = "USB::TEST"
        self.query_responses = query_responses or {}
        self.binary_responses = binary_responses or {}
        self.raw_responses: dict[str, bytes] = {}
        self.serial_log_responses: dict[float, str] = {}
        self.pico_analog_configs: list[object] = []
        self.pico_digital_configs: list[object] = []
        self.saleae_capture_configs: list[object] = []
        self.saleae_uart_configs: list[object] = []
        self.saleae_i2c_configs: list[object] = []
        self.saleae_spi_configs: list[object] = []
        self.saleae_can_configs: list[object] = []
        self.writes: list[str] = []
        self.queries: list[str] = []
        self.binary_queries: list[str] = []
        self.raw_writes: list[str] = []
        self.serial_log_durations: list[float] = []
        self.serial_log_baudrates: list[int | None] = []
        self.serial_commands: list[tuple[str, float, int | None, int | None, str | None, float | None]] = []
        self.ssh_commands: list[tuple[str, str, str, float | None]] = []
        self.serial_configs: list[tuple[int | None, int | None, str | None, float | None]] = []

    def write(self, command: str) -> None:
        self.writes.append(command)

    def query(self, command: str) -> str:
        self.queries.append(command)
        return self.query_responses.get(command, "1.23")

    def query_binary(self, command: str) -> bytes:
        self.binary_queries.append(command)
        return self.binary_responses.get(command, b"DATA")

    def read_raw_after_write(self, command: str, delay_s: float = 0.0) -> bytes:
        self.raw_writes.append(command)
        return self.raw_responses.get(command, b"IN;SP1;PU0,0;PD100,100;")

    def read_serial_log(
        self,
        duration_s: float,
        chunk_timeout_ms: int = 200,
        baudrate: int | None = None,
        bytesize: int | None = None,
        parity: str | None = None,
        stopbits: float | None = None,
        stop_requested=None,
    ) -> str:
        self.serial_log_durations.append(duration_s)
        self.serial_log_baudrates.append(baudrate)
        return self.serial_log_responses.get(duration_s, "serial line 1\nserial line 2")

    def info(self):
        from instrument_visa.visa_client import InstrumentInfo

        return InstrumentInfo(address=self.address, idn=self.query("*IDN?").strip())

    def configure_serial(self, baudrate: int | None = None, bytesize: int | None = None, parity: str | None = None, stopbits: float | None = None) -> None:
        self.serial_configs.append((baudrate, bytesize, parity, stopbits))

    def send_serial_command(
        self,
        command: str,
        response_duration_s: float = 0.0,
        baudrate: int | None = None,
        bytesize: int | None = None,
        parity: str | None = None,
        stopbits: float | None = None,
        stop_requested=None,
    ) -> str:
        self.serial_commands.append((command, response_duration_s, baudrate, bytesize, parity, stopbits))
        return "OK"

    def run_ssh_command(self, command: str, username: str = "", password: str = "", timeout_s: float | None = None, stop_requested=None) -> str:
        self.ssh_commands.append((command, username, password, timeout_s))
        return "ssh output"

    def capture_analog(self, config, stop_requested=None):
        self.pico_analog_configs.append(config)
        return AcquisitionResult(kind="picoscope analog", file_type="csv", content="Time_s,A_V\n0,1.0\n")

    def capture_digital(self, config, stop_requested=None):
        self.pico_digital_configs.append(config)
        return AcquisitionResult(kind="picoscope digital", file_type="csv", content="Time_s,D0\n0,1\n")

    def capture_saleae_digital(self, config, output_dir, stop_requested=None):
        self.saleae_capture_configs.append((config, output_dir))
        return AcquisitionResult(kind="saleae capture", file_type="txt", content="Capture: capture.sal\nRaw CSV: raw")

    def capture_uart(self, config, output_dir, stop_requested=None):
        self.saleae_uart_configs.append((config, output_dir))
        return AcquisitionResult(kind="saleae uart", file_type="txt", content="Capture: capture.sal\nUART CSV: uart.csv")

    def capture_i2c(self, config, output_dir, stop_requested=None):
        self.saleae_i2c_configs.append((config, output_dir))
        return AcquisitionResult(kind="saleae i2c", file_type="txt", content="Capture: capture.sal\nI2C CSV: i2c.csv")

    def capture_spi(self, config, output_dir, stop_requested=None):
        self.saleae_spi_configs.append((config, output_dir))
        return AcquisitionResult(kind="saleae spi", file_type="txt", content="Capture: capture.sal\nSPI CSV: spi.csv")

    def capture_can(self, config, output_dir, stop_requested=None):
        self.saleae_can_configs.append((config, output_dir))
        return AcquisitionResult(kind="saleae can", file_type="txt", content="Capture: capture.sal\nCAN CSV: can.csv")


class AcquisitionTests(unittest.TestCase):
    def test_profile_detection_for_new_manual_checked_devices(self) -> None:
        cases = {
            "KEYSIGHT TECHNOLOGIES,DSOX2024A,MY123,1.0": "keysight_infinivision_x",
            "KEYSIGHT TECHNOLOGIES,MSOX3054T,MY123,1.0": "keysight_infinivision_x",
            "AGILENT TECHNOLOGIES,34401A,MY123,1.0": "keysight_344_l44",
            "AGILENT TECHNOLOGIES,MSO6034A,MY123,1.0": "keysight_infinivision_6000",
            "AGILENT TECHNOLOGIES,DSO7034B,MY123,1.0": "keysight_infinivision_7000",
            "AGILENT TECHNOLOGIES,54622D,MY123,1.0": "agilent_54600",
            "TEKTRONIX,TDS420A,0,CF:91.1CT": "tektronix_tds400",
            "TEKTRONIX,TDS3014B,0,CF:91.1CT": "tektronix_tds30",
            "TEKTRONIX,MDO4104,0,CF:91.1CT": "tektronix_mdo",
            "KEITHLEY INSTRUMENTS INC.,MODEL 2000,123,1.0": "keithley_2000",
            "ROHDE&SCHWARZ,HMS-X,123,1.0": "rs_hameg_hms",
            "LECROY,WAVERUNNER 610ZI,123,1.0": "lecroy_xstream",
            "ROHDE&SCHWARZ,RTB2004,123,1.0": "rs_rt_scope",
            "AGILENT TECHNOLOGIES,4395A,0,1.0": "hp_4395a",
            "HEWLETT-PACKARD,8591A,123,1.0": "hp_8591a",
            "AGILENT TECHNOLOGIES,E7405A,US123,1.0": "hp_e740",
            "AGILENT TECHNOLOGIES,E4402B,US123,1.0": "hp_agilent_e4402b",
            "KEYSIGHT TECHNOLOGIES,N9030A,US123,1.0": "keysight_n90",
            "AGILENT TECHNOLOGIES,E5071C,MY123,1.0": "keysight_e5071c",
            "Rohde&Schwarz,FSW26,123,1.0": "rs_fsw",
            "Rohde&Schwarz,ZNB8,123,1.0": "rs_znb",
            "Rohde&Schwarz,SMIQ03B,123,1.0": "rs_sme_smt_smiq",
            "Rohde&Schwarz,SMHU,123,1.0": "rs_smg_legacy",
            "HAMEG,HMP4030,123,1.0": "rs_hmp_power_supply",
            "HEWLETT-PACKARD,34970A,0,13-2-2": "keysight_34970a",
            "Saleae Logic 2 Automation": "saleae_logic2",
            "PicoScope 2000A": "picoscope",
        }

        for idn, expected_key in cases.items():
            with self.subTest(idn=idn):
                self.assertEqual(detect_profile(idn).key, expected_key)

    def test_keithley_2000_dmm_uses_read_query(self) -> None:
        instrument = FakeInstrument(query_responses={":READ?": "4.56"})

        result = read_value(instrument)  # type: ignore[arg-type]

        self.assertEqual(result.content, "4.56")
        self.assertEqual(instrument.queries, [":READ?"])

    def test_hms_x_trace_uses_csv_format_before_data_query(self) -> None:
        instrument = FakeInstrument(query_responses={"TRAC:DATA?": "[Hz],Trace1[dBm]\n1.0,-2.0"})

        result = capture_waveform(instrument, "ROHDE&SCHWARZ,HMS-X,123,1.0")  # type: ignore[arg-type]

        self.assertEqual(result.file_type, "csv")
        self.assertEqual(instrument.writes, ["TRACe:DATA:FORMat CSV"])
        self.assertEqual(instrument.queries, ["TRAC:DATA?"])
        self.assertIn("[Hz],Trace1[dBm]", str(result.content))

    def test_e4402b_uses_e740_style_trace_export(self) -> None:
        class TraceInstrument(FakeInstrument):
            def query(self, command: str) -> str:
                self.queries.append(command)
                if command.startswith(":MMEM:DATA? 'R:IV"):
                    return "Frequency,Trace\n1.0,-20.0"
                return super().query(command)

        instrument = TraceInstrument()

        result = capture_waveform(instrument, "AGILENT TECHNOLOGIES,E4402B,US123,1.0")  # type: ignore[arg-type]

        self.assertEqual(result.file_type, "csv")
        self.assertEqual(result.content, "Frequency,Trace\n1.0,-20.0")
        self.assertTrue(instrument.writes[0].startswith(":SYST:TIME "))
        self.assertTrue(instrument.writes[1].startswith(":SYST:DATE "))
        self.assertEqual(instrument.writes[2], ":DISP:MENU:STATE 0")
        self.assertRegex(instrument.writes[3], r':MMEM:STOR:TRAC TRACE1,"R:IV[0-9A-F]{8}\.CSV"')
        self.assertRegex(instrument.writes[4], r":MMEM:DEL 'R:IV[0-9A-F]{8}\.CSV'")
        self.assertEqual(instrument.writes[5], ":DISP:MENU:STATE 1")
        self.assertRegex(instrument.queries[0], r":MMEM:DATA\? 'R:IV[0-9A-F]{8}\.CSV'")

    def test_hp_8591a_trace_uses_legacy_trace_query(self) -> None:
        instrument = FakeInstrument(query_responses={"TRA?": "-10,-20,-30"})

        result = capture_waveform(instrument, "HEWLETT-PACKARD,8591A,123,1.0")  # type: ignore[arg-type]

        self.assertEqual(result.file_type, "csv")
        self.assertEqual(result.content, "Point,TraceA\n1,-10\n2,-20\n3,-30")
        self.assertEqual(instrument.queries, ["TRA?"])

    def test_hp_8591a_screenshot_uses_getplot_hpgl(self) -> None:
        instrument = FakeInstrument()
        instrument.raw_responses["GETPLOT"] = b"IN;SP1;PU0,0;PD100,100;"

        result = capture_screenshot(instrument, "HEWLETT-PACKARD,8591A,123,1.0")  # type: ignore[arg-type]

        self.assertEqual(result.file_type, "hpgl")
        self.assertEqual(result.content, b"IN;SP1;PU0,0;PD100,100;")
        self.assertEqual(instrument.raw_writes, ["GETPLOT"])

    def test_signal_generator_read_uses_basic_scpi_queries(self) -> None:
        instrument = FakeInstrument(query_responses={":SOUR:FREQ:CW?": "100000000", ":SOUR:POW?": "-30", ":OUTP?": "1"})

        settings = read_signal_generator_settings(instrument, "Rohde&Schwarz,SMIQ03B,123,1.0")  # type: ignore[arg-type]

        self.assertEqual(settings.frequency, "100000000")
        self.assertEqual(settings.power, "-30")
        self.assertEqual(settings.rf_output, "ON")
        self.assertEqual(instrument.queries, [":SOUR:FREQ:CW?", ":SOUR:POW?", ":OUTP?"])

    def test_signal_generator_set_switches_rf_off_before_change(self) -> None:
        instrument = FakeInstrument(query_responses={":SOUR:FREQ:CW?": "100000000", ":SOUR:POW?": "-30", ":OUTP?": "0"})

        result = set_signal_generator(instrument, "Rohde&Schwarz,SMIQ03B,123,1.0", "100 MHz", "-30 dBm", False, 0.0, True)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes, [":OUTP OFF", ":SOUR:FREQ:CW 100 MHz", ":SOUR:POW -30 dBm", ":OUTP OFF"])
        self.assertIn("Frequency,100000000", str(result.content))

    def test_signal_generator_rejects_power_above_limit(self) -> None:
        instrument = FakeInstrument()

        with self.assertRaises(ValueError):
            set_signal_generator(instrument, "Rohde&Schwarz,SMIQ03B,123,1.0", "100 MHz", "10 dBm", True, 0.0, True)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes, [])

    def test_signal_generator_rejects_injected_frequency_before_writes(self) -> None:
        instrument = FakeInstrument()

        with self.assertRaises(ValueError):
            set_signal_generator(instrument, "Rohde&Schwarz,SMIQ03B,123,1.0", "100 MHz;:OUTP ON", "-30 dBm", True, 0.0, True)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes, [])

    def test_signal_generator_rf_off_command(self) -> None:
        instrument = FakeInstrument(query_responses={":SOUR:FREQ:CW?": "100000000", ":SOUR:POW?": "-30", ":OUTP?": "0"})

        result = set_signal_generator_rf_output(instrument, "Rohde&Schwarz,SMIQ03B,123,1.0", False)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes, [":OUTP OFF"])
        self.assertIn("RFOutput,OFF", str(result.content))

    def test_smg_legacy_generator_uses_manual_iec_commands(self) -> None:
        instrument = FakeInstrument(query_responses={"RF?": "RF 123456000", "LEVEL:RF?": "LEVEL:RF -20"})

        result = set_signal_generator(instrument, "Rohde&Schwarz,SMHU,123,1.0", "123.456MHz", "-20DBM", True, 0.0, True)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes, ["LEVEL:RF:OFF", "RF 123.456MHz", "LEVEL:RF -20DBM", "LEVEL:RF:ON"])
        self.assertEqual(instrument.queries, ["RF?", "LEVEL:RF?"])
        self.assertIn("Frequency,123456000", str(result.content))
        self.assertIn("Power,-20", str(result.content))
        self.assertIn("RFOutput,ON", str(result.content))

    def test_smg_legacy_generator_reads_off_state(self) -> None:
        instrument = FakeInstrument(query_responses={"RF?": "RF 123456000", "LEVEL:RF?": "LEVEL:RF:OFF"})

        settings = read_signal_generator_settings(instrument, "Rohde&Schwarz,SMGU,123,1.0")  # type: ignore[arg-type]

        self.assertEqual(settings.frequency, "123456000")
        self.assertEqual(settings.power, "OFF")
        self.assertEqual(settings.rf_output, "OFF")

    def test_smg_legacy_generator_accepts_alternate_off_responses(self) -> None:
        for level_response in ("OFF", "0", "LEVEL:RF OFF", "LEVEL:RF:0"):
            with self.subTest(level_response=level_response):
                instrument = FakeInstrument(query_responses={"RF?": "RF 123456000", "LEVEL:RF?": level_response})

                settings = read_signal_generator_settings(instrument, "Rohde&Schwarz,SMHU,123,1.0")  # type: ignore[arg-type]

                self.assertEqual(settings.rf_output, "OFF")

    def test_hmp4030_power_supply_read_uses_selected_channel(self) -> None:
        instrument = FakeInstrument(
            query_responses={
                "VOLT?": "5.000",
                "CURR?": "0.500",
                "MEAS:VOLT?": "4.998",
                "MEAS:CURR?": "0.123",
                "OUTP:SEL?": "1",
                "OUTP:GEN?": "1",
            }
        )

        settings = read_power_supply_settings(instrument, "HAMEG,HMP4030,123,1.0", 2)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes, ["INST:NSEL 2"])
        self.assertEqual(instrument.queries, ["VOLT?", "CURR?", "MEAS:VOLT?", "MEAS:CURR?", "OUTP:SEL?", "OUTP:GEN?"])
        self.assertEqual(settings.voltage_set, "5.000")
        self.assertEqual(settings.output_selected, "ON")

    def test_hmp4030_power_supply_set_uses_safe_channel_sequence(self) -> None:
        instrument = FakeInstrument(
            query_responses={
                "VOLT?": "5.000",
                "CURR?": "0.500",
                "MEAS:VOLT?": "4.998",
                "MEAS:CURR?": "0.123",
                "OUTP:SEL?": "1",
                "OUTP:GEN?": "1",
            }
        )

        result = set_power_supply(instrument, "HAMEG,HMP4030,123,1.0", 1, "5 V", "0.5 A", True, 10, 1)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes[:5], ["INST:NSEL 1", "VOLT 5", "CURR 0.5", "OUTP:SEL 1", "OUTP:GEN 1"])
        self.assertIn("VoltageMeasured,4.998", str(result.content))

    def test_hmp4030_rejects_voltage_above_limit_before_writes(self) -> None:
        instrument = FakeInstrument()

        with self.assertRaises(ValueError):
            set_power_supply(instrument, "HAMEG,HMP4030,123,1.0", 1, "20 V", "0.5 A", True, 10, 1)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes, [])

    def test_hmp4030_rejects_invalid_channel_before_writes(self) -> None:
        instrument = FakeInstrument()

        with self.assertRaises(ValueError):
            set_power_supply(instrument, "HAMEG,HMP4030,123,1.0", 4, "5 V", "0.5 A", True, 10, 1)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes, [])

    def test_hmp2020_rejects_channel_three_before_writes(self) -> None:
        instrument = FakeInstrument()

        with self.assertRaises(ValueError):
            read_power_supply_settings(instrument, "HAMEG,HMP2020,123,1.0", 3)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes, [])

    def test_hmp4030_rejects_negative_voltage_and_current(self) -> None:
        instrument = FakeInstrument()

        with self.assertRaises(ValueError):
            set_power_supply(instrument, "HAMEG,HMP4030,123,1.0", 1, "-1 V", "0.5 A", True, 10, 1)  # type: ignore[arg-type]
        with self.assertRaises(ValueError):
            set_power_supply(instrument, "HAMEG,HMP4030,123,1.0", 1, "5 V", "-0.5 A", True, 10, 1)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes, [])

    def test_hmp4030_power_supply_output_off(self) -> None:
        instrument = FakeInstrument(
            query_responses={
                "VOLT?": "5.000",
                "CURR?": "0.500",
                "MEAS:VOLT?": "0.000",
                "MEAS:CURR?": "0.000",
                "OUTP:SEL?": "0",
                "OUTP:GEN?": "1",
            }
        )

        result = set_power_supply_output(instrument, "HAMEG,HMP4030,123,1.0", 3, False)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes[:2], ["INST:NSEL 3", "OUTP:SEL 0"])
        self.assertIn("OutputSelected,OFF", str(result.content))

    def test_hmp4030_power_supply_master_output_off(self) -> None:
        instrument = FakeInstrument(
            query_responses={
                "VOLT?": "5.000",
                "CURR?": "0.500",
                "MEAS:VOLT?": "0.000",
                "MEAS:CURR?": "0.000",
                "OUTP:SEL?": "0",
                "OUTP:GEN?": "0",
            }
        )

        result = set_power_supply_master_output(instrument, "HAMEG,HMP4030,123,1.0", False, 1)  # type: ignore[arg-type]

        self.assertEqual(instrument.writes[:2], ["OUTP:GEN 0", "INST:NSEL 1"])
        self.assertIn("OutputGeneral,OFF", str(result.content))

    def test_hms_x_screenshot_normalizes_ieee_bmp_block(self) -> None:
        payload = b"#6000010BM12345678"
        instrument = FakeInstrument(binary_responses={"HCOPy:DATA?": payload})

        result = capture_screenshot(instrument, "ROHDE&SCHWARZ,HMS-X,123,1.0")  # type: ignore[arg-type]

        self.assertEqual(result.file_type, "bmp")
        self.assertEqual(result.content, b"BM12345678")
        self.assertEqual(instrument.writes, ["HCOPy:FORMat BMP"])
        self.assertEqual(instrument.binary_queries, ["HCOPy:DATA?"])

    def test_tektronix_tds400_measurement_commands_are_manual_style(self) -> None:
        instrument = FakeInstrument(query_responses={"MEASUrement:IMMed:VALue?": "1.0"})

        result = read_scope_measurement(instrument, "Vrms", 2, "TEKTRONIX,TDS420A,0,CF:91.1CT")  # type: ignore[arg-type]

        self.assertEqual(result.content, "1.0")
        self.assertEqual(instrument.writes, ["MEASUrement:IMMed:SOUrce CH2", "MEASUrement:IMMed:TYPe RMS"])
        self.assertEqual(instrument.queries, ["MEASUrement:IMMed:VALue?"])

    def test_rs_rt_scope_measurement_commands_are_manual_style(self) -> None:
        instrument = FakeInstrument(query_responses={"MEASurement1:RESult:ACTual?": "2.0"})

        result = read_scope_measurement(instrument, "Vpp", 1, "ROHDE&SCHWARZ,RTB2004,123,1.0")  # type: ignore[arg-type]

        self.assertEqual(result.content, "2.0")
        self.assertEqual(instrument.writes, ["MEASurement1:SOURce C1", "MEASurement1:MAIN PEAK"])
        self.assertEqual(instrument.queries, ["MEASurement1:RESult:ACTual?"])

    def test_lecroy_screenshot_normalizes_bmp_after_prefix(self) -> None:
        instrument = FakeInstrument()
        instrument.raw_responses["SCREEN_DUMP"] = b"LECROY HEADER BMabcdef"

        result = capture_screenshot(instrument, "LECROY,WAVERUNNER 610ZI,123,1.0")  # type: ignore[arg-type]

        self.assertEqual(result.file_type, "bmp")
        self.assertEqual(result.content, b"BMabcdef")
        self.assertEqual(instrument.writes, ["HARDCOPY_SETUP DEV,BMP,PORT,GPIB"])
        self.assertEqual(instrument.raw_writes, ["SCREEN_DUMP"])

    def test_e5071c_screenshot_restores_display_on_failure(self) -> None:
        class FailingBinaryInstrument(FakeInstrument):
            def query_binary(self, command: str) -> bytes:
                self.binary_queries.append(command)
                raise RuntimeError("binary failed")

        instrument = FailingBinaryInstrument(query_responses={":DISP:MENU:STATE?": "1", ":DISP:IMAG?": "NORM", ":DISP:SKEY:STAT?": "1"})

        with self.assertRaises(RuntimeError):
            capture_screenshot(instrument, "AGILENT TECHNOLOGIES,E5071C,MY123,1.0")  # type: ignore[arg-type]

        self.assertEqual(instrument.writes[-3:], [":DISP:IMAG NORM", ":DISP:SKEY:STAT 1", ":DISP:MENU:STATE 1"])

    def test_infinivision_7000_screenshot_waits_after_query_trigger(self) -> None:
        class RawInstrument(FakeInstrument):
            def __init__(self) -> None:
                super().__init__()
                self.raw_calls: list[tuple[str, float]] = []

            def read_raw_after_write(self, command: str, delay_s: float = 0.0) -> bytes:
                self.raw_calls.append((command, delay_s))
                return b"\x89PNG\r\n\x1a\nDATA"

        instrument = RawInstrument()

        result = capture_screenshot(instrument, "AGILENT TECHNOLOGIES,DSO7034B,MY123,1.0")  # type: ignore[arg-type]

        self.assertEqual(result.file_type, "png")
        self.assertEqual(instrument.writes, [":HARD:INKS ON"])
        self.assertEqual(instrument.raw_calls, [(":DISP:DATA? PNG, SCR, COL", 2.0)])

    def test_e740_screenshot_rejects_injected_title_before_title_write(self) -> None:
        instrument = FakeInstrument()

        with self.assertRaises(ValueError):
            capture_screenshot(instrument, "AGILENT TECHNOLOGIES,E7405A,US123,1.0", "bad';:INST:DEL")  # type: ignore[arg-type]

        self.assertTrue(instrument.writes[0].startswith(":SYST:TIME "))
        self.assertTrue(instrument.writes[1].startswith(":SYST:DATE "))
        self.assertEqual(instrument.writes[2:], [])

    def test_e740_screenshot_uses_unique_temp_filename(self) -> None:
        instrument = FakeInstrument(binary_responses={})

        capture_screenshot(instrument, "AGILENT TECHNOLOGIES,E7405A,US123,1.0")  # type: ignore[arg-type]

        self.assertNotIn("R:INTUI.WMF", "\n".join([*instrument.writes, *instrument.binary_queries]))
        self.assertRegex(instrument.writes[-2], r":MMEM:DEL 'R:IV[0-9A-F]{8}\.WMF'")

    def test_keysight_e5071c_sparameters_use_unique_temp_filename(self) -> None:
        instrument = FakeInstrument(query_responses={":MMEM:TRAN? 'D:\\KILO12345678.s1p';*WAI": "! data"})

        capture_sparameters(instrument, "AGILENT TECHNOLOGIES,E5071C,MY123,1.0", SParameterConfig(format="DB", s1=True))  # type: ignore[arg-type]

        joined = "\n".join([*instrument.writes, *instrument.queries])
        self.assertNotIn("SNP01", joined)
        self.assertRegex(joined, r"D:\\KILO[0-9A-F]{8}\.s1p")

    def test_sparameters_rejects_invalid_format_before_writes(self) -> None:
        instrument = FakeInstrument()

        with self.assertRaises(ValueError):
            capture_sparameters(instrument, "AGILENT TECHNOLOGIES,E5071C,MY123,1.0", SParameterConfig(format="DB;*RST", s1=True))  # type: ignore[arg-type]

        self.assertEqual(instrument.writes, [])

    def test_znb_sparameters_rejects_port_above_reported_count(self) -> None:
        instrument = FakeInstrument(query_responses={"INST:NSEL?": "1", "SOUR:GRO:PPOR?": "1,2"})

        with self.assertRaises(ValueError):
            capture_sparameters(instrument, "Rohde&Schwarz,ZNB8,123,1.0", SParameterConfig(format="DB", s4=True))  # type: ignore[arg-type]

        self.assertEqual(instrument.writes, [])

    def test_znb_sparameters_use_unique_temp_filename(self) -> None:
        instrument = FakeInstrument(query_responses={"INST:NSEL?": "1", "SOUR:GRO:PPOR?": "1,2", "MMEM:DATA? 'Traces\\kilo12345678.s1p'": "! data"})

        capture_sparameters(instrument, "Rohde&Schwarz,ZNB8,123,1.0", SParameterConfig(format="DB", s1=True))  # type: ignore[arg-type]

        joined = "\n".join([*instrument.writes, *instrument.queries])
        self.assertNotIn("visatmp", joined)
        self.assertRegex(joined, r"Traces\\KILO[0-9A-F]{8}\.s1p")

    def test_znb_screenshot_can_capture_full_page(self) -> None:
        instrument = FakeInstrument()

        capture_screenshot(instrument, "Rohde&Schwarz,ZNB8,123,1.0", znb_full_page=True)  # type: ignore[arg-type]

        self.assertIn("HCOP:PAGE:WIND ALL", instrument.writes)
        self.assertNotIn("HCOP:PAGE:WIND ACT", instrument.writes)

    def test_waveform_export_adds_excel_chart(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "results.xlsx"
            result = AcquisitionResult(kind="waveform", file_type="csv", content="Time,CH1,CH2\n0,1,2\n1,3,4\n")

            export = append_result(workbook_path, "USB::TEST", "TEST,DSOX2024A,1,1", result)
            workbook = load_workbook(export.workbook_path)
            sheet = workbook[export.sheet_name]

            self.assertEqual(sheet["A7"].value, "Time")
            self.assertEqual(sheet["B8"].value, 1.0)
            self.assertEqual(len(sheet._charts), 1)
            self.assertEqual(sheet.freeze_panes, "A8")

    def test_excel_export_escapes_formula_like_values(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "results.xlsx"
            result = AcquisitionResult(kind="serial log", file_type="csv", content="Name,Value\nDUT,=HYPERLINK(\"http://example.invalid\")\n")

            export = append_result(workbook_path, "COM1", "Serial", result)
            workbook = load_workbook(export.workbook_path, data_only=False)
            sheet = workbook[export.sheet_name]

            self.assertEqual(sheet["B8"].value, "'=HYPERLINK(\"http://example.invalid\")")

    def test_excel_export_uses_results_sheet_even_when_other_sheet_active(self) -> None:
        from openpyxl import Workbook, load_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "results.xlsx"
            workbook = Workbook()
            results = workbook.active
            results.title = "Results"
            results.append(["Timestamp", "Address", "IDN", "Kind", "FileType", "ValueOrFile"])
            data = workbook.create_sheet("Waveform")
            data.append(["do not append here"])
            workbook.active = workbook.sheetnames.index("Waveform")
            workbook.save(workbook_path)

            append_result(workbook_path, "USB::TEST", "IDN", AcquisitionResult(kind="value", file_type="value", content="1.23"))
            loaded = load_workbook(workbook_path)

            self.assertEqual(loaded["Results"].max_row, 2)
            self.assertEqual(loaded["Waveform"].max_row, 1)

    def test_excel_export_keeps_nan_inf_as_text(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "results.xlsx"
            result = AcquisitionResult(kind="value table", file_type="csv", content="Name,Value\nA,NaN\nB,Inf\nC,-Inf\nD,1.5\n")

            export = append_result(workbook_path, "USB::TEST", "IDN", result)
            workbook = load_workbook(export.workbook_path)
            sheet = workbook[export.sheet_name]

            self.assertEqual(sheet["B8"].value, "NaN")
            self.assertEqual(sheet["B9"].value, "Inf")
            self.assertEqual(sheet["B10"].value, "'-Inf")
            self.assertEqual(sheet["B11"].value, 1.5)

    def test_large_waveform_export_adds_min_max_extract_for_chart(self) -> None:
        from openpyxl import load_workbook

        rows = ["Time,CH1"]
        for index in range(2500):
            value = 100 if index == 1999 else index % 10
            rows.append(f"{index},{value}")

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "results.xlsx"
            result = AcquisitionResult(kind="waveform", file_type="csv", content="\n".join(rows))

            export = append_result(workbook_path, "USB::TEST", "TEST,DSOX2024A,1,1", result)
            workbook = load_workbook(export.workbook_path)
            sheet = workbook[export.sheet_name]

            self.assertIn("Diagramm-Extrakt: Min/Max", str(sheet["D6"].value))
            self.assertEqual(sheet["D7"].value, "Time")
            self.assertEqual(sheet["E7"].value, "CH1")
            self.assertEqual(sheet["E5"].value, True)
            self.assertTrue(any("B2007" in str(sheet.cell(row, 5).value) for row in range(8, sheet.max_row + 1)))
            self.assertEqual(len(sheet._charts), 1)

    def test_frequency_parsing_and_points(self) -> None:
        self.assertEqual(parse_frequency_hz("100 MHz"), 100_000_000.0)
        self.assertEqual(frequency_points("100 MHz", "102 MHz", "1 MHz"), [100_000_000.0, 101_000_000.0, 102_000_000.0])

    def test_frequency_points_rejects_unbounded_sweeps(self) -> None:
        with self.assertRaises(ValueError):
            frequency_points("1 Hz", f"{MAX_SWEEP_POINTS + 1} Hz", "1 Hz")

    def test_voltage_points(self) -> None:
        self.assertEqual(voltage_points("0 V", "2 V", "1 V"), [0.0, 1.0, 2.0])

    def test_frequency_sweep_sets_generator_and_reads_dmm(self) -> None:
        generator = FakeInstrument(query_responses={
            "*IDN?": "Rohde&Schwarz,SMIQ03B,123,1.0",
            ":SOUR:FREQ:CW?": "100000000",
            ":SOUR:POW?": "-30",
            ":OUTP?": "1",
        })
        generator.address = "GPIB0::1::INSTR"
        measurement = FakeInstrument(query_responses={"*IDN?": "KEITHLEY INSTRUMENTS INC.,MODEL 2000,123,1.0", ":READ?": "4.2"})
        measurement.address = "GPIB0::2::INSTR"

        result = run_frequency_sweep(
            generator,  # type: ignore[arg-type]
            measurement,  # type: ignore[arg-type]
            FrequencySweepConfig(
                start_frequency="100 MHz",
                stop_frequency="101 MHz",
                step_frequency="1 MHz",
                power="-30 dBm",
                max_power_dbm=0,
                settle_s=0,
                measurement_mode="dmm",
            ),
        )

        self.assertEqual(result.actual_count, 2)
        self.assertEqual(result.ok_count, 2)
        self.assertIn("SetFrequencyHz,Index", result.csv_content)
        self.assertIn("100000000.000000", result.csv_content)
        self.assertIn("101000000.000000", result.csv_content)
        self.assertEqual(generator.writes[-1], ":OUTP OFF")
        self.assertEqual(measurement.queries, ["*IDN?", ":READ?", ":READ?"])

    def test_custom_sequence_repeats_variable_and_reads_dmm(self) -> None:
        generator = FakeInstrument(query_responses={
            "*IDN?": "Rohde&Schwarz,SMIQ03B,123,1.0",
            ":SOUR:FREQ:CW?": "100000000",
            ":SOUR:POW?": "-30",
            ":OUTP?": "1",
        })
        generator.address = "GPIB0::1::INSTR"
        dmm = FakeInstrument(query_responses={"*IDN?": "KEITHLEY INSTRUMENTS INC.,MODEL 2000,123,1.0", ":READ?": "4.2"})
        dmm.address = "GPIB0::2::INSTR"

        result = run_custom_sequence(
            {"Generator1": generator, "DMM1": dmm},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Generator1": generator.address, "DMM1": dmm.address},
                steps=[
                    SequenceStep("Generator1", "generator_set_frequency", {"frequency": "${frequency}", "power": "-30 dBm", "max_power_dbm": "0", "rf": "ON"}),
                    SequenceStep("DMM1", "dmm_read"),
                ],
                repeat=2,
                variables=[SequenceVariable("frequency", "100 MHz", "1 MHz", "frequency")],
            ),
        )

        self.assertEqual(result.actual_count, 4)
        self.assertEqual(result.ok_count, 4)
        self.assertIn(":SOUR:FREQ:CW 100000000HZ", generator.writes)
        self.assertIn(":SOUR:FREQ:CW 101000000HZ", generator.writes)
        self.assertEqual(dmm.queries, ["*IDN?", ":READ?", ":READ?"])
        self.assertIn("generator_set_frequency", result.csv_content)

    def test_custom_sequence_can_capture_spectrum_trace_summary(self) -> None:
        analyzer = FakeInstrument(query_responses={"*IDN?": "HEWLETT-PACKARD,8591A,123,1.0", "TRA?": "1,2,3"})
        analyzer.address = "GPIB0::18::INSTR"

        result = run_custom_sequence(
            {"Spectrum1": analyzer},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Spectrum1": analyzer.address},
                steps=[SequenceStep("Spectrum1", "capture_waveform", {"channels": "", "point_mode": "RAW"})],
            ),
        )

        self.assertEqual(result.ok_count, 1)
        self.assertIn("waveform", result.csv_content)
        self.assertEqual(analyzer.queries, ["*IDN?", "TRA?"])

    def test_custom_sequence_exports_screenshot_step_artifact(self) -> None:
        instrument = FakeInstrument(query_responses={"*IDN?": "TEST,DSOX2024A,1,1"}, binary_responses={":DISPLAY:DATA? PNG, COLOR": b"PNGDATA"})
        exports: list[tuple[str, str, str, bytes | str]] = []

        def export_step(device: str, info, result: AcquisitionResult) -> str:
            exports.append((device, info.address, result.file_type, result.content))
            return "Datei: screenshot.png"

        result = run_custom_sequence(
            {"Scope1": instrument},  # type: ignore[dict-item]
            CustomSequenceConfig(devices={"Scope1": instrument.address}, steps=[SequenceStep("Scope1", "capture_screenshot")], end_rf_off=False),
            step_result_export=export_step,
        )

        self.assertEqual(result.ok_count, 1)
        self.assertEqual(exports, [("Scope1", "USB::TEST", "png", b"PNGDATA")])
        self.assertIn("Datei: screenshot.png", result.csv_content)

    def test_custom_sequence_znb_screenshot_full_page_param(self) -> None:
        instrument = FakeInstrument(query_responses={"*IDN?": "Rohde&Schwarz,ZNB8,123,1.0"})

        result = run_custom_sequence(
            {"ZNB1": instrument},  # type: ignore[dict-item]
            CustomSequenceConfig(devices={"ZNB1": instrument.address}, steps=[SequenceStep("ZNB1", "capture_screenshot", {"znb_full_page": "ON"})], end_rf_off=False),
        )

        self.assertEqual(result.ok_count, 1)
        self.assertIn("HCOP:PAGE:WIND ALL", instrument.writes)

    def test_msox3054t_waveform_uses_3000x_scpi(self) -> None:
        instrument = FakeInstrument(
            query_responses={
                ":CHANnel1:DISPlay?": "1",
                ":WAVeform:DATA?": "1.0,2.0",
                ":SYSTem:ERRor?": "+0,No error",
                ":WAVeform:XINCrement?": "0.5",
                ":WAVeform:XREFerence?": "0",
                ":WAVeform:XORigin?": "0",
                ":SYSTEM:ERROR?": "+0,No error",
            }
        )

        result = capture_waveform(instrument, "KEYSIGHT TECHNOLOGIES,MSOX3054T,MY123,1.0", channels=[1], point_mode="RAW")

        self.assertEqual(result.file_type, "csv")
        self.assertIn("CH1", result.content)
        self.assertIn("0.0,1.0", result.content)
        self.assertIn("0.5,2.0", result.content)
        self.assertEqual(
            instrument.writes,
            [":WAVeform:SOURce CHANnel1", ":WAVeform:FORMat ASCII", ":WAVeform:POINts:MODE RAW"],
        )
        self.assertEqual(
            instrument.queries,
            [
                ":CHANnel1:DISPlay?",
                ":WAVeform:DATA?",
                ":SYSTem:ERRor?",
                ":WAVeform:XINCrement?",
                ":WAVeform:XREFerence?",
                ":WAVeform:XORigin?",
                ":SYSTEM:ERROR?",
            ],
        )

    def test_custom_sequence_exports_serial_log_step(self) -> None:
        instrument = FakeInstrument(query_responses={"*IDN?": "SERIAL,LOGGER,1,1"})
        instrument.address = "ASRL3::INSTR"
        exports: list[tuple[str, str, str, bytes | str]] = []

        def export_step(device: str, info, result: AcquisitionResult) -> str:
            exports.append((device, info.address, result.file_type, result.content))
            return "Datei: serial-log.txt"

        result = run_custom_sequence(
            {"Serial1": instrument},  # type: ignore[dict-item]
            CustomSequenceConfig(devices={"Serial1": instrument.address}, steps=[SequenceStep("Serial1", "serial_log", {"duration_s": "2.5", "baudrate": "115200", "serial_format": "8N1"})], end_rf_off=False),
            step_result_export=export_step,
        )

        self.assertEqual(result.ok_count, 1)
        self.assertEqual(instrument.serial_log_durations, [2.5])
        self.assertEqual(instrument.serial_log_baudrates, [115200])
        self.assertEqual(exports, [("Serial1", "ASRL3::INSTR", "txt", "serial line 1\nserial line 2")])
        self.assertIn("Datei: serial-log.txt", result.csv_content)
        self.assertNotIn("serial line 1\nserial line 2", result.csv_content)

    def test_custom_sequence_serial_log_device_does_not_require_idn(self) -> None:
        instrument = FakeInstrument()
        instrument.address = "ASRL4::INSTR"

        def failing_query(command: str) -> str:
            raise TimeoutError(command)

        instrument.query = failing_query  # type: ignore[method-assign]

        result = run_custom_sequence(
            {"Serial1": instrument},  # type: ignore[dict-item]
            CustomSequenceConfig(devices={"Serial1": instrument.address}, steps=[SequenceStep("Serial1", "serial_log", {"duration_s": "1", "baudrate": "9600", "serial_format": "8N1"})], end_rf_off=False),
        )

        self.assertEqual(result.ok_count, 1)
        self.assertEqual(instrument.queries, [])
        self.assertIn("Serielles Gerät ohne IDN", result.csv_content)
        self.assertEqual(instrument.serial_log_durations, [1.0])

    def test_custom_sequence_sends_serial_command(self) -> None:
        instrument = FakeInstrument()
        instrument.address = "ASRL5::INSTR"

        result = run_custom_sequence(
            {"Serial1": instrument},  # type: ignore[dict-item]
            CustomSequenceConfig(devices={"Serial1": instrument.address}, steps=[SequenceStep("Serial1", "serial_command", {"command": "PING\\n", "baudrate": "57600", "serial_format": "7E1", "response_duration_s": "0.5"})], end_rf_off=False),
        )

        self.assertEqual(result.ok_count, 1)
        self.assertEqual(instrument.serial_commands, [("PING\\n", 0.5, 57600, 7, "E", 1)])
        self.assertIn("OK", result.csv_content)

    def test_custom_sequence_serial_command_device_does_not_require_idn(self) -> None:
        instrument = FakeInstrument()
        instrument.address = "ASRL6::INSTR"

        def failing_query(command: str) -> str:
            raise TimeoutError(command)

        instrument.query = failing_query  # type: ignore[method-assign]

        result = run_custom_sequence(
            {"Serial1": instrument},  # type: ignore[dict-item]
            CustomSequenceConfig(devices={"Serial1": instrument.address}, steps=[SequenceStep("Serial1", "serial_command", {"command": "BOOT\\r", "baudrate": "9600", "serial_format": "8N1"})], end_rf_off=False),
        )

        self.assertEqual(result.ok_count, 1)
        self.assertEqual(instrument.queries, [])
        self.assertIn("Serielles Gerät ohne IDN", result.csv_content)

    def test_custom_sequence_runs_ssh_command(self) -> None:
        instrument = FakeInstrument(query_responses={"*IDN?": "SSH test"})
        instrument.address = "ssh://tester@example.local:2222"

        result = run_custom_sequence(
            {"SSH1": instrument},  # type: ignore[dict-item]
            CustomSequenceConfig(devices={"SSH1": instrument.address}, steps=[SequenceStep("SSH1", "ssh_command", {"command": "uname -a", "username": "override", "password": "secret", "timeout_s": "12"})], end_rf_off=False),
        )

        self.assertEqual(result.ok_count, 1)
        self.assertEqual(instrument.ssh_commands, [("uname -a", "override", "secret", 12.0)])
        self.assertIn("ssh output", result.csv_content)
        self.assertIn("password=***", result.csv_content)
        self.assertNotIn("password=secret", result.csv_content)

    def test_parse_ssh_target_supports_user_host_port(self) -> None:
        target = parse_ssh_target("ssh://user@example.local:2222")

        self.assertEqual(target.host, "example.local")
        self.assertEqual(target.port, 2222)
        self.assertEqual(target.username, "user")

    def test_run_ssh_command_uses_paramiko(self) -> None:
        class FakeChannel:
            def __init__(self) -> None:
                self.stdout_reads = [b"hello\n"]
                self.stderr_reads: list[bytes] = []

            def exit_status_ready(self) -> bool:
                return True

            def recv_ready(self) -> bool:
                return bool(self.stdout_reads)

            def recv_stderr_ready(self) -> bool:
                return bool(self.stderr_reads)

            def recv(self, size: int) -> bytes:
                return self.stdout_reads.pop(0)

            def recv_stderr(self, size: int) -> bytes:
                return self.stderr_reads.pop(0)

            def recv_exit_status(self) -> int:
                return 0

            def close(self) -> None:
                return None

        class FakeStdout:
            def __init__(self) -> None:
                self.channel = FakeChannel()

        class FakeClient:
            connect_kwargs: dict[str, object] = {}
            command = ""
            closed = False
            loaded_host_keys = False
            missing_policy = None

            def load_system_host_keys(self) -> None:
                self.__class__.loaded_host_keys = True

            def set_missing_host_key_policy(self, policy: object) -> None:
                self.__class__.missing_policy = policy
                return None

            def connect(self, **kwargs: object) -> None:
                self.__class__.connect_kwargs = kwargs

            def exec_command(self, command: str, timeout: float):
                self.__class__.command = command
                return None, FakeStdout(), SimpleNamespace(channel=FakeChannel())

            def close(self) -> None:
                self.__class__.closed = True

        reject_policy = object()
        fake_paramiko = SimpleNamespace(SSHClient=FakeClient, RejectPolicy=lambda: reject_policy)
        import instrument_visa.sequence as sequence_module

        original_paramiko = sequence_module.paramiko
        sequence_module.paramiko = fake_paramiko  # type: ignore[assignment]
        try:
            result = run_ssh_command("ssh://user@example.local:2222", "echo hello", timeout_s=5)
        finally:
            sequence_module.paramiko = original_paramiko  # type: ignore[assignment]

        self.assertEqual(result, "hello")
        self.assertEqual(FakeClient.command, "echo hello")
        self.assertEqual(FakeClient.connect_kwargs["hostname"], "example.local")
        self.assertEqual(FakeClient.connect_kwargs["port"], 2222)
        self.assertEqual(FakeClient.connect_kwargs["username"], "user")
        self.assertTrue(FakeClient.loaded_host_keys)
        self.assertIs(FakeClient.missing_policy, reject_policy)
        self.assertTrue(FakeClient.closed)

    def test_read_direct_serial_log_decodes_bytes(self) -> None:
        class FakeSerial:
            def __init__(self, port: str, baudrate: int, bytesize: int, parity: str, stopbits: float, timeout: float) -> None:
                self.port = port
                self.baudrate = baudrate
                self.bytesize = bytesize
                self.parity = parity
                self.stopbits = stopbits
                self.timeout = timeout
                self.reads = [b"boot ", b"ok\n", b""]

            def __enter__(self):
                return self

            def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
                return None

            @property
            def in_waiting(self) -> int:
                return len(self.reads[0]) if self.reads else 0

            def read(self, size: int) -> bytes:
                if self.reads:
                    return self.reads.pop(0)
                return b""

        import instrument_visa.sequence as sequence_module

        original_serial = sequence_module.serial
        sequence_module.serial = SimpleNamespace(Serial=FakeSerial)  # type: ignore[assignment]
        try:
            result = read_direct_serial_log("COM3", 0.01, 115200, 7, "E", 1)
        finally:
            sequence_module.serial = original_serial  # type: ignore[assignment]

        self.assertEqual(result, "boot ok\n")

    def test_send_direct_serial_command_writes_escaped_bytes_and_reads_response(self) -> None:
        class FakeSerial:
            writes: list[bytes] = []

            def __init__(self, port: str, baudrate: int, bytesize: int, parity: str, stopbits: float, timeout: float, write_timeout: float) -> None:
                self.port = port
                self.baudrate = baudrate
                self.bytesize = bytesize
                self.parity = parity
                self.stopbits = stopbits
                self.timeout = timeout
                self.write_timeout = write_timeout
                self.reads = [b"ACK\r\n", b""]

            def __enter__(self):
                return self

            def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
                return None

            @property
            def in_waiting(self) -> int:
                return len(self.reads[0]) if self.reads else 0

            def write(self, data: bytes) -> None:
                self.writes.append(data)

            def flush(self) -> None:
                return None

            def read(self, size: int) -> bytes:
                if self.reads:
                    return self.reads.pop(0)
                return b""

        import instrument_visa.sequence as sequence_module

        original_serial = sequence_module.serial
        sequence_module.serial = SimpleNamespace(Serial=FakeSerial)  # type: ignore[assignment]
        try:
            result = send_direct_serial_command("COM7", "PING\\r\\n", 115200, response_duration_s=0.01)
        finally:
            sequence_module.serial = original_serial  # type: ignore[assignment]

        self.assertEqual(result, "ACK\r\n")
        self.assertEqual(FakeSerial.writes, [b"PING\r\n"])

    def test_probe_direct_serial_idn_returns_idn_and_settings(self) -> None:
        class FakeSerial:
            writes: list[bytes] = []

            def __init__(self, port: str, baudrate: int, bytesize: int, parity: str, stopbits: float, timeout: float, write_timeout: float, xonxoff: bool, rtscts: bool, dsrdtr: bool) -> None:
                self.port = port
                self.baudrate = baudrate
                self.bytesize = bytesize
                self.parity = parity
                self.stopbits = stopbits
                self.timeout = timeout
                self.write_timeout = write_timeout
                self.xonxoff = xonxoff
                self.rtscts = rtscts
                self.dsrdtr = dsrdtr
                self.reads = [b"*IDN?\r\nTEST,MODEL,123,1\r\n", b""]

            def __enter__(self):
                return self

            def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
                return None

            @property
            def in_waiting(self) -> int:
                return len(self.reads[0]) if self.reads else 0

            def reset_input_buffer(self) -> None:
                return None

            def reset_output_buffer(self) -> None:
                return None

            def write(self, data: bytes) -> None:
                self.writes.append(data)

            def read(self, size: int) -> bytes:
                if self.reads:
                    return self.reads.pop(0)
                return b""

        import instrument_visa.sequence as sequence_module

        original_serial = sequence_module.serial
        original_settings = list(sequence_module._SERIAL_SCPI_PREFERRED_SETTINGS)
        original_cached_settings = dict(sequence_module._SERIAL_SCPI_SETTINGS)
        sequence_module._SERIAL_SCPI_PREFERRED_SETTINGS.clear()
        sequence_module._SERIAL_SCPI_SETTINGS.clear()
        sequence_module.serial = SimpleNamespace(Serial=FakeSerial)  # type: ignore[assignment]
        try:
            idn, settings = probe_direct_serial_idn("COM3", 500)
        finally:
            sequence_module.serial = original_serial  # type: ignore[assignment]
            sequence_module._SERIAL_SCPI_PREFERRED_SETTINGS[:] = original_settings
            sequence_module._SERIAL_SCPI_SETTINGS.clear()
            sequence_module._SERIAL_SCPI_SETTINGS.update(original_cached_settings)

        self.assertEqual(idn, "TEST,MODEL,123,1")
        self.assertEqual(settings, (9600, "8N1", "none", "\n"))
        self.assertEqual(FakeSerial.writes, [b"*IDN?\n"])

    def test_probe_direct_serial_idn_detects_34970a_with_short_probe(self) -> None:
        class FakeSerial:
            writes: list[tuple[int, bytes]] = []

            def __init__(self, port: str, baudrate: int, bytesize: int, parity: str, stopbits: float, timeout: float, write_timeout: float, xonxoff: bool, rtscts: bool, dsrdtr: bool) -> None:
                self.port = port
                self.baudrate = baudrate
                self.bytesize = bytesize
                self.parity = parity
                self.stopbits = stopbits
                self.timeout = timeout
                self.write_timeout = write_timeout
                self.xonxoff = xonxoff
                self.rtscts = rtscts
                self.dsrdtr = dsrdtr
                self.reads = [b"HEWLETT-PACKARD,34970A,0,13-2-2\r\n", b""] if baudrate == 19200 else [b""]

            def __enter__(self):
                return self

            def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
                return None

            @property
            def in_waiting(self) -> int:
                return len(self.reads[0]) if self.reads else 0

            def reset_input_buffer(self) -> None:
                return None

            def reset_output_buffer(self) -> None:
                return None

            def write(self, data: bytes) -> None:
                self.writes.append((self.baudrate, data))

            def read(self, size: int) -> bytes:
                if self.reads:
                    return self.reads.pop(0)
                return b""

        import instrument_visa.sequence as sequence_module

        original_serial = sequence_module.serial
        original_settings = list(sequence_module._SERIAL_SCPI_PREFERRED_SETTINGS)
        original_cached_settings = dict(sequence_module._SERIAL_SCPI_SETTINGS)
        sequence_module._SERIAL_SCPI_PREFERRED_SETTINGS.clear()
        sequence_module._SERIAL_SCPI_SETTINGS.clear()
        sequence_module.serial = SimpleNamespace(Serial=FakeSerial)  # type: ignore[assignment]
        try:
            idn, settings = probe_direct_serial_idn("COM7", 500, exhaustive=False)
        finally:
            sequence_module.serial = original_serial  # type: ignore[assignment]
            sequence_module._SERIAL_SCPI_PREFERRED_SETTINGS[:] = original_settings
            sequence_module._SERIAL_SCPI_SETTINGS.clear()
            sequence_module._SERIAL_SCPI_SETTINGS.update(original_cached_settings)

        self.assertEqual(idn, "HEWLETT-PACKARD,34970A,0,13-2-2")
        self.assertEqual(detect_profile(idn).key, "keysight_34970a")
        self.assertEqual(settings, (19200, "8N1", "none", "\n"))

    def test_parse_serial_format(self) -> None:
        self.assertEqual(parse_serial_format("8N1"), (8, "N", 1.0))
        self.assertEqual(parse_serial_format("7-E-1"), (7, "E", 1.0))
        self.assertEqual(parse_serial_format("8N1.5"), (8, "N", 1.5))

        with self.assertRaises(ValueError):
            parse_serial_format("9N1")

    def test_parse_parallel_tasks(self) -> None:
        tasks = parse_parallel_tasks("DMM1:dmm; Scope1:scope:Vpp:2; Serial1:serial:9600:8N1")

        self.assertEqual([task.device for task in tasks], ["DMM1", "Scope1", "Serial1"])
        self.assertEqual(tasks[1].measurement, "Vpp")
        self.assertEqual(tasks[1].channel, 2)
        self.assertEqual(tasks[2].baudrate, 9600)

    def test_custom_sequence_parallel_phase_runs_measurements(self) -> None:
        dmm = FakeInstrument(query_responses={"*IDN?": "KEITHLEY INSTRUMENTS INC.,MODEL 2000,123,1.0", ":READ?": "4.2"})
        dmm.address = "GPIB0::2::INSTR"
        scope = FakeInstrument(query_responses={"*IDN?": "TEST,DSOX2024A,1,1", ":MEASure:VPP? CHANnel1": "1.1"})
        scope.address = "USB::SCOPE"
        exports: list[tuple[str, str, str]] = []

        def export_step(device: str, info, result: AcquisitionResult) -> str:
            exports.append((device, info.address, result.file_type))
            return "Tabellenblatt: parallel"

        result = run_custom_sequence(
            {"DMM1": dmm, "Scope1": scope},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"DMM1": dmm.address, "Scope1": scope.address},
                steps=[SequenceStep("", "parallel_phase", {"duration_s": "0.01", "interval_s": "0.01", "tasks": "DMM1:dmm; Scope1:scope:Vpp:1"})],
                end_rf_off=False,
            ),
            step_result_export=export_step,
        )

        self.assertEqual(result.ok_count, 1)
        self.assertIn("parallel phase", result.csv_content)
        self.assertEqual(exports[0][2], "csv")

    def test_append_result_writes_serial_log_text_artifact(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "results.xlsx"
            export = append_result(workbook_path, "ASRL3::INSTR", "Serieller Log", AcquisitionResult(kind="serial log", file_type="txt", content="line 1\nline 2"))

            self.assertTrue(workbook_path.exists())
            self.assertIsNotNone(export.artifact_path)
            self.assertEqual(export.artifact_path.read_text(encoding="utf-8"), "line 1\nline 2")

    def test_append_result_serializes_concurrent_workbook_writes(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "results.xlsx"
            result = AcquisitionResult(kind="value", file_type="value", content="1.23")

            with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
                list(executor.map(lambda index: append_result(workbook_path, f"USB::{index}", "TEST", result), range(12)))

            workbook = load_workbook(workbook_path)
            self.assertEqual(workbook.active.max_row, 13)
            addresses = [row[1] for row in workbook.active.iter_rows(min_row=2, values_only=True)]
            self.assertEqual(set(addresses), {f"USB::{index}" for index in range(12)})

    def test_picoscope_channel_parsers(self) -> None:
        self.assertEqual(parse_pico_analog_channels("A,B,CHC"), ["A", "B", "C"])
        self.assertEqual(parse_pico_digital_channels("D0-D3,D7"), [0, 1, 2, 3, 7])

        with self.assertRaises(ValueError):
            parse_pico_analog_channels("Z")
        with self.assertRaises(ValueError):
            parse_pico_digital_channels("D16")
        with self.assertRaises(ValueError):
            parse_pico_digital_channels("D7-D0")

    def test_picoscope_variant_capabilities(self) -> None:
        self.assertEqual(picoscope_analog_channels_for_variant("2206BMSO"), ("A", "B"))
        self.assertEqual(picoscope_analog_channels_for_variant("2206B MSO"), ("A", "B"))
        self.assertEqual(picoscope_analog_channels_for_variant("2406B"), ("A", "B", "C", "D"))
        self.assertTrue(picoscope_variant_supports_digital("2206BMSO"))
        self.assertTrue(picoscope_variant_supports_digital("2206B MSO"))
        self.assertFalse(picoscope_variant_supports_digital("2406B"))
        self.assertFalse(picoscope_variant_supports_digital("unknown MSO"))
        with self.assertRaises(ValueError):
            picoscope_analog_channels_for_variant("")
        with self.assertRaises(ValueError):
            picoscope_analog_channels_for_variant("unknown")

    def test_custom_sequence_picoscope_analog_step(self) -> None:
        pico = FakeInstrument(query_responses={"*IDN?": "PicoScope 2000A"})
        pico.address = "PICO2000A::AUTO"
        exports: list[tuple[str, str]] = []

        def export_step(device: str, info, result: AcquisitionResult) -> str:
            exports.append((device, result.file_type))
            return "Tabellenblatt: PicoAnalog"

        result = run_custom_sequence(
            {"Pico1": pico},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Pico1": pico.address},
                steps=[SequenceStep("Pico1", "picoscope_analog", {"channels": "A,B", "range": "5V", "samples": "100", "interval_us": "2"})],
                end_rf_off=False,
            ),
            step_result_export=export_step,
        )

        self.assertEqual(result.ok_count, 1)
        self.assertEqual(pico.pico_analog_configs[0].channels, "A,B")
        self.assertEqual(exports, [("Pico1", "csv")])

    def test_custom_sequence_picoscope_digital_step(self) -> None:
        pico = FakeInstrument(query_responses={"*IDN?": "PicoScope 2000A"})
        pico.address = "PICO2000A::AUTO"

        result = run_custom_sequence(
            {"Pico1": pico},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Pico1": pico.address},
                steps=[SequenceStep("Pico1", "picoscope_digital", {"channels": "D0-D7", "logic_level_mv": "1500", "samples": "100", "interval_us": "2"})],
                end_rf_off=False,
            ),
        )

        self.assertEqual(result.ok_count, 1)
        self.assertEqual(pico.pico_digital_configs[0].channels, "D0-D7")

    def test_parse_34970a_channels(self) -> None:
        self.assertEqual(parse_34970a_channels("1-3,5"), [1, 2, 3, 5])

        with self.assertRaises(ValueError):
            parse_34970a_channels("0")
        with self.assertRaises(ValueError):
            parse_34970a_channels("23")

    def test_34970a_voltage_read_uses_channel_list(self) -> None:
        instrument = FakeInstrument(query_responses={"MEAS:VOLT:DC? 10,0.003,(@101,102,103)": "1.0,2.0,3.0"})

        result = read_34970a_data_logger(instrument, DataLogger34970AConfig(measurement="VOLT_DC", channels="1-3"))

        self.assertEqual(result.file_type, "csv")
        self.assertIn("101", instrument.queries[0])
        self.assertIn("Timestamp,CH1 VOLT_DC [V],CH2 VOLT_DC [V],CH3 VOLT_DC [V]", str(result.content))
        self.assertIn(",1.0,2.0,3.0", str(result.content))

    def test_34970a_voltage_read_uses_single_query_for_large_channel_lists(self) -> None:
        instrument = FakeInstrument(
            query_responses={
                "MEAS:VOLT:DC? 10,0.003,(@101,102,103,104,105,106,107,108,109,110,111,112)": "1,2,3,4,5,6,7,8,9,10,11,12",
            }
        )

        result = read_34970a_data_logger(instrument, DataLogger34970AConfig(measurement="VOLT_DC", channels="1-12"))

        self.assertEqual(len(instrument.queries), 1)
        self.assertIn("CH12 VOLT_DC [V]", str(result.content))
        self.assertIn(",1,2,3,4,5,6,7,8,9,10,11,12", str(result.content))

    def test_34970a_temperature_read_sets_serial_and_tc_type(self) -> None:
        instrument = FakeInstrument(query_responses={"MEAS:TEMP? TC,K,(@101)": "23.5"})

        result = read_34970a_data_logger(instrument, DataLogger34970AConfig(measurement="TEMP", channels="1", thermocouple_type="K", baudrate=19200, serial_format="7E1"))

        self.assertIn("Timestamp,CH1 TEMP [degC]", str(result.content))
        self.assertIn(",23.5", str(result.content))
        self.assertEqual(instrument.serial_configs, [(19200, 7, "E", 1.0)])

    def test_34970a_rejects_injected_numeric_params_before_measurement(self) -> None:
        instrument = FakeInstrument()

        with self.assertRaises(ValueError):
            read_34970a_data_logger(instrument, DataLogger34970AConfig(measurement="VOLT_DC", channels="1", range_value="10;*RST"))

        self.assertEqual(instrument.queries, [])

    def test_34970a_rejects_invalid_thermocouple_type(self) -> None:
        instrument = FakeInstrument()

        with self.assertRaises(ValueError):
            read_34970a_data_logger(instrument, DataLogger34970AConfig(measurement="TEMP", channels="1", thermocouple_type="K;*RST"))

        self.assertEqual(instrument.queries, [])

    def test_34970a_excel_export_appends_wide_rows_to_same_sheet(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "results.xlsx"
            first = AcquisitionResult(kind="34970A measurement plan", file_type="csv", content="Timestamp,CH1 TEMP [degC],CH21 CURR_DC [A]\n2026-07-08 14:00:00,23.5,0.1\n")
            second = AcquisitionResult(kind="34970A measurement plan", file_type="csv", content="Timestamp,CH1 TEMP [degC],CH21 CURR_DC [A]\n2026-07-08 14:00:05,23.6,0.2\n")

            first_export = append_result(workbook_path, "COM7", "HEWLETT-PACKARD,34970A", first)
            second_export = append_result(workbook_path, "COM7", "HEWLETT-PACKARD,34970A", second)
            workbook = load_workbook(workbook_path)
            sheet = workbook["34970A Measurements"]

            self.assertEqual(first_export.sheet_name, "34970A Measurements")
            self.assertEqual(second_export.sheet_name, "34970A Measurements")
            self.assertEqual(sheet.max_row, 3)
            self.assertEqual(sheet["A1"].value, "Timestamp")
            self.assertEqual(sheet["B1"].value, "CH1 TEMP [degC]")
            self.assertEqual(sheet["A2"].value, "2026-07-08 14:00:00")
            self.assertEqual(sheet["B3"].value, 23.6)

    def test_34970a_excel_export_adds_late_stopped_header(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "results.xlsx"
            first = AcquisitionResult(kind="34970A data logger", file_type="csv", content="Timestamp,CH1 TEMP [degC]\n2026-07-08 14:00:00,23.5\n")
            stopped = AcquisitionResult(kind="34970A data logger", file_type="csv", content="Timestamp,StoppedByUser\n2026-07-08 14:00:05,Yes\n")

            append_result(workbook_path, "COM7", "HEWLETT-PACKARD,34970A", first)
            append_result(workbook_path, "COM7", "HEWLETT-PACKARD,34970A", stopped)
            workbook = load_workbook(workbook_path)
            sheet = workbook["34970A Measurements"]

            self.assertEqual(sheet["C1"].value, "StoppedByUser")
            self.assertEqual(sheet["C3"].value, "Yes")

    def test_parse_ca410_measurement_response(self) -> None:
        values = parse_ca410_measurement_response("OK00,P1,0,0.3274345,0.4191236,4.8075729,+0.39,2.1047971,1.0,2.0,3.0")

        self.assertEqual(values["Status"], "OK00")
        self.assertEqual(values["Probe"], "P1")
        self.assertEqual(values["x"], "0.3274345")
        self.assertEqual(values["y"], "0.4191236")
        self.assertEqual(values["Lv"], "4.8075729")
        self.assertEqual(values["FMAFlickerPercent"], "2.1047971")
        self.assertEqual(values["Z"], "3.0")

    def test_ca410_read_sets_serial_and_exports_xy_lv(self) -> None:
        instrument = FakeInstrument(
            query_responses={
                "COM,1": "OK00",
                "OPR,1": "OK00",
                "SCS,3": "OK00",
                "VSN,2": "OK00",
                "FSC,1": "OK00",
                "MMS,0": "OK00",
                "FMS,0": "OK00",
                "MCH,1,0": "OK00",
                "MDS,0": "OK00",
                "MES,2": "OK00,P1,0,0.3274345,0.4191236,4.8075729,+0.39,2.1047971,1.0,2.0,3.0",
            }
        )

        result = read_ca410_measurement(instrument, CA410Config())

        self.assertEqual(instrument.serial_configs, [(38400, 7, "E", 2.0)])
        self.assertEqual(instrument.queries, ["COM,1", "OPR,1", "SCS,3", "FSC,1", "MMS,0", "VSN,2", "FMS,0", "MCH,1,0", "MDS,0", "MES,2"])
        self.assertIn("Timestamp,Status,Probe,DisplayMode,x,y,Lv,TempShift,FMAFlickerPercent,X,Y,Z", str(result.content))
        self.assertIn("OK00,P1,0,0.3274345,0.4191236,4.8075729,+0.39,2.1047971,1.0,2.0,3.0", str(result.content))

    def test_custom_sequence_ca410_step(self) -> None:
        instrument = FakeInstrument(
            query_responses={
                "*IDN?": "Konica Minolta CA-410",
                "COM,1": "OK00",
                "OPR,1": "OK00",
                "SCS,4,60.00": "OK00",
                "VSN,1": "OK00",
                "FSC,1": "OK00",
                "MMS,1": "OK00",
                "FMS,0": "OK00",
                "MCH,1,2": "OK00",
                "MDS,1": "OK00",
                "MES,2": "OK00,P1,1,6500.0,0.001,120.5,+0.01,0.5,1.0,2.0,3.0",
            }
        )
        instrument.address = "COM9"

        result = run_custom_sequence(
            {"CA410": instrument},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"CA410": instrument.address},
                steps=[SequenceStep("CA410", "ca410_read", {"color_mode": "TcpduvLv", "probe": "1", "calibration_channel": "2", "measurement_method": "Color", "sync_mode": "INT", "sync_value": "60", "integration_mode": "Single-Frame", "baudrate": "38400", "serial_format": "7E2"})],
                end_rf_off=False,
            ),
        )

        self.assertEqual(result.ok_count, 1)
        self.assertIn("CA-410 measurement", result.csv_content)
        self.assertIn("MCH,1,2", instrument.queries)
        self.assertIn("SCS,4,60.00", instrument.queries)
        self.assertIn("VSN,1", instrument.queries)
        self.assertIn("MDS,1", instrument.queries)

    def test_ca410_read_averages_for_configured_time(self) -> None:
        instrument = FakeInstrument(
            query_responses={
                "COM,1": "OK00",
                "OPR,1": "OK00",
                "SCS,3": "OK00",
                "VSN,2": "OK00",
                "FSC,1": "OK00",
                "MMS,1": "OK00",
                "FMS,0": "OK00",
                "MCH,1,0": "OK00",
                "MDS,0": "OK00",
                "MES,2": "OK00,P1,0,0.2000000,0.4000000,10.000000,+0.00,0.0,1.0,2.0,3.0",
            }
        )

        result = read_ca410_measurement(instrument, CA410Config(measurement_method="Color", averaging_time_s=0.01))

        self.assertIn("AverageSamples", str(result.content))
        self.assertIn("0.2,0.4,10", str(result.content))

    def test_ca410_single_frame_requires_synchronous_mode(self) -> None:
        instrument = FakeInstrument()

        with self.assertRaisesRegex(ValueError, "Sync"):
            read_ca410_measurement(instrument, CA410Config(measurement_method="Color", integration_mode="Single-Frame"))

    def test_ca410_excel_export_appends_rows(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "results.xlsx"
            first = AcquisitionResult(kind="CA-410 measurement", file_type="csv", content="Timestamp,Status,Probe,DisplayMode,x,y,Lv\n2026-07-13 14:00:00,OK00,P1,0,0.1,0.2,100\n")
            second = AcquisitionResult(kind="CA-410 measurement", file_type="csv", content="Timestamp,Status,Probe,DisplayMode,x,y,Lv\n2026-07-13 14:00:01,OK00,P1,0,0.3,0.4,101\n")

            first_export = append_result(workbook_path, "COM9", "Konica Minolta CA-410", first)
            second_export = append_result(workbook_path, "COM9", "Konica Minolta CA-410", second)
            workbook = load_workbook(workbook_path)
            sheet = workbook["CA-410 Measurements"]

            self.assertEqual(first_export.sheet_name, "CA-410 Measurements")
            self.assertEqual(second_export.sheet_name, "CA-410 Measurements")
            self.assertEqual(sheet.max_row, 3)
            self.assertEqual(sheet["B1"].value, "Status")
            self.assertEqual(sheet["E3"].value, 0.3)

    def test_parse_34970a_measurement_plan(self) -> None:
        tasks = parse_34970a_measurement_plan("1-4:VOLT_DC; 5:TEMP; 6-7:RES")

        self.assertEqual([(task.channels, task.measurement) for task in tasks], [("1-4", "VOLT_DC"), ("5", "TEMP"), ("6-7", "RES")])

        with self.assertRaises(ValueError):
            parse_34970a_measurement_plan("1-4")

    def test_custom_sequence_34970a_step(self) -> None:
        instrument = FakeInstrument(query_responses={"*IDN?": "HEWLETT-PACKARD,34970A,0,1", "MEAS:RES? (@101,102)": "10,20"})
        instrument.address = "COM5"

        result = run_custom_sequence(
            {"Logger1": instrument},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Logger1": instrument.address},
                steps=[SequenceStep("Logger1", "data_logger_34970a_read", {"measurement": "RES", "channels": "1-2", "range": "AUTO", "resolution": "DEF", "thermocouple_type": "K", "baudrate": "19200", "serial_format": "8N1"})],
                end_rf_off=False,
            ),
        )

        self.assertEqual(result.ok_count, 1)
        self.assertIn("34970A data logger", result.csv_content)

    def test_custom_sequence_34970a_measurement_plan_step(self) -> None:
        instrument = FakeInstrument(
            query_responses={
                "*IDN?": "HEWLETT-PACKARD,34970A,0,1",
                "MEAS:VOLT:DC? 10,0.003,(@101,102)": "1,2",
                "MEAS:TEMP? TC,K,(@103)": "23",
            }
        )

        result = run_custom_sequence(
            {"Logger1": instrument},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Logger1": "COM5"},
                steps=[SequenceStep("Logger1", "data_logger_34970a_plan", {"plan": "1-2:VOLT_DC; 3:TEMP", "baudrate": "19200", "serial_format": "8N1"})],
                end_rf_off=False,
            ),
        )

        self.assertEqual(result.ok_count, 1)
        self.assertIn("34970A measurement plan", result.csv_content)

    def test_parse_saleae_channels(self) -> None:
        self.assertEqual(parse_saleae_channels("D0-D3,D7"), [0, 1, 2, 3, 7])

        with self.assertRaises(ValueError):
            parse_saleae_channels("D16")
        with self.assertRaises(ValueError):
            parse_saleae_channels("-1,D0")

    def test_bool_param_rejects_invalid_text(self) -> None:
        self.assertTrue(_bool_param({"rf": "ON"}, "rf", False))
        self.assertFalse(_bool_param({"rf": "nein"}, "rf", True))
        with self.assertRaises(ValueError):
            _bool_param({"rf": "treu"}, "rf", False)

    def test_unknown_profile_does_not_advertise_unsupported_features(self) -> None:
        profile = detect_profile("ACME,MODEL,344-SERIAL,1.0")

        self.assertEqual(profile.key, "unknown")
        self.assertFalse(profile.supports_screenshot)
        self.assertFalse(profile.supports_waveform)
        self.assertFalse(profile.supports_dmm_read)

    def test_custom_sequence_saleae_capture_step(self) -> None:
        saleae = FakeInstrument(query_responses={"*IDN?": "Saleae"})
        saleae.address = "SALEAE::LOCAL"

        result = run_custom_sequence(
            {"Saleae1": saleae},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Saleae1": saleae.address},
                steps=[SequenceStep("Saleae1", "saleae_capture", {"channels": "D0-D1", "duration_s": "1", "sample_rate": "1000000", "threshold_v": "3.3"})],
                end_rf_off=False,
            ),
        )

        self.assertEqual(result.ok_count, 1)
        config, output_dir = saleae.saleae_capture_configs[0]
        self.assertIsInstance(config, SaleaeCaptureConfig)
        self.assertEqual(config.digital_channels, "D0-D1")
        self.assertEqual(str(output_dir), "saleae_output")

    def test_custom_sequence_saleae_uart_step(self) -> None:
        saleae = FakeInstrument(query_responses={"*IDN?": "Saleae"})
        saleae.address = "SALEAE::LOCAL"

        result = run_custom_sequence(
            {"Saleae1": saleae},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Saleae1": saleae.address},
                steps=[SequenceStep("Saleae1", "saleae_uart", {"channel": "0", "baudrate": "115200", "duration_s": "1", "sample_rate": "1000000", "threshold_v": "3.3"})],
                end_rf_off=False,
            ),
        )

        self.assertEqual(result.ok_count, 1)
        config, output_dir = saleae.saleae_uart_configs[0]
        self.assertIsInstance(config, SaleaeUartConfig)
        self.assertEqual(config.channel, 0)
        self.assertEqual(str(output_dir), "saleae_output")

    def test_custom_sequence_saleae_i2c_step(self) -> None:
        saleae = FakeInstrument(query_responses={"*IDN?": "Saleae"})
        saleae.address = "SALEAE::LOCAL"

        result = run_custom_sequence(
            {"Saleae1": saleae},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Saleae1": saleae.address},
                steps=[SequenceStep("Saleae1", "saleae_i2c", {"sda": "0", "scl": "1", "duration_s": "1", "sample_rate": "1000000"})],
                end_rf_off=False,
            ),
        )

        self.assertEqual(result.ok_count, 1)
        config, output_dir = saleae.saleae_i2c_configs[0]
        self.assertIsInstance(config, SaleaeI2cConfig)
        self.assertEqual(config.sda_channel, 0)
        self.assertEqual(config.scl_channel, 1)

    def test_custom_sequence_saleae_spi_step(self) -> None:
        saleae = FakeInstrument(query_responses={"*IDN?": "Saleae"})
        saleae.address = "SALEAE::LOCAL"

        result = run_custom_sequence(
            {"Saleae1": saleae},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Saleae1": saleae.address},
                steps=[SequenceStep("Saleae1", "saleae_spi", {"mosi": "0", "miso": "1", "clock": "2", "duration_s": "1"})],
                end_rf_off=False,
            ),
        )

        self.assertEqual(result.ok_count, 1)
        config, output_dir = saleae.saleae_spi_configs[0]
        self.assertIsInstance(config, SaleaeSpiConfig)
        self.assertEqual(config.clock_channel, 2)
        self.assertEqual(config.enable_channel, -1)

    def test_saleae_spi_capture_without_enable_uses_only_required_channels(self) -> None:
        from instrument_visa.saleae_client import SaleaeInstrument

        captured: dict[str, object] = {}

        class FakeAutomation:
            class LogicDeviceConfiguration:
                def __init__(self, **kwargs: object) -> None:
                    captured["channels"] = kwargs["enabled_digital_channels"]

            class CaptureConfiguration:
                def __init__(self, **kwargs: object) -> None:
                    return None

            class TimedCaptureMode:
                def __init__(self, duration_seconds: float) -> None:
                    return None

        class FakeCapture:
            def __enter__(self):
                return self

            def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
                return None

            def wait(self) -> None:
                return None

            def add_analyzer(self, *args: object, **kwargs: object) -> object:
                return object()

            def export_data_table(self, **kwargs: object) -> None:
                return None

            def save_capture(self, **kwargs: object) -> None:
                return None

        class FakeManager:
            def __enter__(self):
                return self

            def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
                return None

            def start_capture(self, **kwargs: object):
                return FakeCapture()

        import instrument_visa.saleae_client as saleae_module

        original_connect = saleae_module._connect_saleae_manager
        original_automation = saleae_module._saleae_automation
        saleae_module._connect_saleae_manager = lambda: FakeManager()  # type: ignore[assignment]
        saleae_module._saleae_automation = lambda: FakeAutomation  # type: ignore[assignment]
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                SaleaeInstrument().capture_spi(SaleaeSpiConfig(0, 1, 2, -1, 0.01, 1000000), Path(temp_dir))
        finally:
            saleae_module._connect_saleae_manager = original_connect  # type: ignore[assignment]
            saleae_module._saleae_automation = original_automation  # type: ignore[assignment]

        self.assertEqual(captured["channels"], [0, 1, 2])

    def test_sanitized_step_params_redacts_passwords(self) -> None:
        self.assertEqual(sanitized_step_params({"command": "id", "password": "secret"}), {"command": "id", "password": "***"})

    def test_custom_sequence_saleae_can_step(self) -> None:
        saleae = FakeInstrument(query_responses={"*IDN?": "Saleae"})
        saleae.address = "SALEAE::LOCAL"

        result = run_custom_sequence(
            {"Saleae1": saleae},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Saleae1": saleae.address},
                steps=[SequenceStep("Saleae1", "saleae_can", {"channel": "0", "bitrate": "500000", "duration_s": "1", "sample_rate": "1000000"})],
                end_rf_off=False,
            ),
        )

        self.assertEqual(result.ok_count, 1)
        config, output_dir = saleae.saleae_can_configs[0]
        self.assertIsInstance(config, SaleaeCanConfig)
        self.assertEqual(config.bitrate, 500000)

    def test_custom_sequence_power_supply_uses_configured_safety_limits(self) -> None:
        supply = FakeInstrument(query_responses={"*IDN?": "HAMEG,HMP4030,123,1.0"})

        result = run_custom_sequence(
            {"Supply1": supply},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Supply1": supply.address},
                steps=[SequenceStep("Supply1", "power_supply_set", {"voltage": "10 V", "current": "0.1 A", "channel": "1", "output": "ON"})],
                power_supply_max_voltage=5,
                power_supply_max_current=1,
            )
        )

        self.assertEqual(result.error_count, 1)
        self.assertIn("exceeds max voltage", result.csv_content)
        self.assertEqual(supply.writes, [])

    def test_custom_sequence_validation_rejects_bad_imported_step(self) -> None:
        instrument = FakeInstrument(query_responses={"*IDN?": "TEST,UNKNOWN,1,1"})

        with self.assertRaises(ValueError):
            run_custom_sequence(
                {"Device1": instrument},  # type: ignore[dict-item]
                CustomSequenceConfig(devices={"Device1": instrument.address}, steps=[SequenceStep("Device1", "unknown_action")]),
            )

    def test_cli_loads_json_sequence_config(self) -> None:
        content = {
            "devices": {"DMM1": "GPIB0::2::INSTR"},
            "repeat": 2,
            "pause_s": 0,
            "variables": [{"name": "value", "start": "0", "step": "1", "unit": "number"}],
            "steps": [{"device": "DMM1", "action": "dmm_read", "params": {}}],
            "end_rf_off": False,
            "end_power_supply_off": False,
            "power_supply_max_voltage": 5,
            "power_supply_max_current": 1,
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            sequence_path = Path(temp_dir) / "ablauf.json"
            sequence_path.write_text(__import__("json").dumps(content), encoding="utf-8")

            config = _load_sequence_config(sequence_path)

        self.assertEqual(config.devices, {"DMM1": "GPIB0::2::INSTR"})
        self.assertEqual(config.repeat, 2)
        self.assertEqual(config.steps, [SequenceStep("DMM1", "dmm_read", {})])
        self.assertEqual(config.variables[0].name, "value")
        self.assertFalse(config.end_rf_off)

    def test_cli_sequence_config_parses_boolean_strings(self) -> None:
        content = {
            "devices": {"DMM1": "GPIB0::2::INSTR"},
            "steps": [{"device": "DMM1", "action": "dmm_read", "params": {}}],
            "end_rf_off": "false",
            "end_power_supply_off": "true",
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            sequence_path = Path(temp_dir) / "ablauf.json"
            sequence_path.write_text(__import__("json").dumps(content), encoding="utf-8")

            config = _load_sequence_config(sequence_path)

        self.assertFalse(config.end_rf_off)
        self.assertTrue(config.end_power_supply_off)

    def test_cli_sequence_config_rejects_invalid_boolean_strings(self) -> None:
        with self.assertRaises(ValueError):
            parse_json_bool("maybe", True)

    def test_cli_list_all_resources_includes_non_visa_devices(self) -> None:
        import instrument_visa.cli as cli_module

        original_list_resources = cli_module.list_resources
        original_list_direct_serial_ports = cli_module.list_direct_serial_ports
        original_list_picoscope_resources = cli_module.list_picoscope_resources
        original_list_saleae_resources = cli_module.list_saleae_resources
        cli_module.list_resources = lambda: ["USB::1::INSTR", "COM3"]  # type: ignore[assignment]
        cli_module.list_direct_serial_ports = lambda: [SimpleNamespace(device="COM3"), SimpleNamespace(device="COM4")]  # type: ignore[assignment]
        cli_module.list_picoscope_resources = lambda: ["PICO2000A::AUTO"]  # type: ignore[assignment]
        cli_module.list_saleae_resources = lambda: ["SALEAE::LOCAL"]  # type: ignore[assignment]
        try:
            resources = _list_all_resources()
        finally:
            cli_module.list_resources = original_list_resources  # type: ignore[assignment]
            cli_module.list_direct_serial_ports = original_list_direct_serial_ports  # type: ignore[assignment]
            cli_module.list_picoscope_resources = original_list_picoscope_resources  # type: ignore[assignment]
            cli_module.list_saleae_resources = original_list_saleae_resources  # type: ignore[assignment]

        self.assertEqual(resources, ["USB::1::INSTR", "COM3", "COM4", "PICO2000A::AUTO", "SALEAE::LOCAL"])

    def test_visa_open_closes_resource_manager_if_open_resource_fails(self) -> None:
        import instrument_visa.visa_client as visa_module

        class FakeResourceManager:
            closed = False

            def open_resource(self, address: str) -> object:
                raise RuntimeError(address)

            def close(self) -> None:
                self.__class__.closed = True

        original_resource_manager = visa_module.pyvisa.ResourceManager
        visa_module.pyvisa.ResourceManager = lambda: FakeResourceManager()  # type: ignore[assignment]
        try:
            with self.assertRaises(RuntimeError):
                VisaInstrument("USB::FAIL").open()
        finally:
            visa_module.pyvisa.ResourceManager = original_resource_manager  # type: ignore[assignment]

        self.assertTrue(FakeResourceManager.closed)

    def test_visa_close_closes_resource_manager_if_instrument_close_fails(self) -> None:
        class FakeInstrument:
            def close(self) -> None:
                raise RuntimeError("instrument close failed")

        class FakeResourceManager:
            closed = False

            def close(self) -> None:
                self.__class__.closed = True

        instrument = VisaInstrument("USB::TEST")
        instrument._instrument = FakeInstrument()  # type: ignore[attr-defined]
        instrument._resource_manager = FakeResourceManager()  # type: ignore[attr-defined]

        with self.assertRaises(RuntimeError):
            instrument.close()

        self.assertTrue(FakeResourceManager.closed)
        self.assertIsNone(instrument._instrument)  # type: ignore[attr-defined]
        self.assertIsNone(instrument._resource_manager)  # type: ignore[attr-defined]

    def test_custom_sequence_stops_on_step_error_and_cleans_up_generator(self) -> None:
        generator = FakeInstrument(query_responses={"*IDN?": "Rohde&Schwarz,SMIQ03B,123,1.0", ":SOUR:FREQ:CW?": "100000000", ":SOUR:POW?": "-30", ":OUTP?": "1"})

        def failing_write(command: str) -> None:
            if command == ":SOUR:FREQ:CW 100 MHz":
                raise RuntimeError("generator failed")
            generator.writes.append(command)

        generator.write = failing_write  # type: ignore[method-assign]
        result = run_custom_sequence(
            {"Generator1": generator},  # type: ignore[dict-item]
            CustomSequenceConfig(
                devices={"Generator1": generator.address},
                steps=[SequenceStep("Generator1", "generator_set_frequency", {"frequency": "100 MHz", "power": "-30 dBm", "max_power_dbm": "0"})],
            ),
        )

        self.assertEqual(result.error_count, 1)
        self.assertIn(":OUTP OFF", generator.writes)
        self.assertIn("generator failed", result.csv_content)

    def test_frequency_sweep_rejects_power_above_limit_before_writes(self) -> None:
        generator = FakeInstrument(query_responses={"*IDN?": "Rohde&Schwarz,SMIQ03B,123,1.0"})
        measurement = FakeInstrument(query_responses={"*IDN?": "KEITHLEY INSTRUMENTS INC.,MODEL 2000,123,1.0"})

        with self.assertRaises(ValueError):
            run_frequency_sweep(
                generator,  # type: ignore[arg-type]
                measurement,  # type: ignore[arg-type]
                FrequencySweepConfig(
                    start_frequency="100 MHz",
                    stop_frequency="101 MHz",
                    step_frequency="1 MHz",
                    power="10 dBm",
                    max_power_dbm=0,
                    settle_s=0,
                    measurement_mode="dmm",
                ),
            )

        self.assertEqual(generator.writes, [])

    def test_frequency_sweep_can_stop_before_first_point(self) -> None:
        generator = FakeInstrument(query_responses={"*IDN?": "Rohde&Schwarz,SMIQ03B,123,1.0", ":SOUR:FREQ:CW?": "100000000", ":SOUR:POW?": "-30", ":OUTP?": "0"})
        measurement = FakeInstrument(query_responses={"*IDN?": "KEITHLEY INSTRUMENTS INC.,MODEL 2000,123,1.0"})

        result = run_frequency_sweep(
            generator,  # type: ignore[arg-type]
            measurement,  # type: ignore[arg-type]
            FrequencySweepConfig("100 MHz", "101 MHz", "1 MHz", "-30 dBm", 0, 0, "dmm"),
            stop_requested=lambda: True,
        )

        self.assertEqual(result.actual_count, 0)
        self.assertTrue(result.stopped)
        self.assertIn("StoppedByUser,Yes", result.csv_content)

    def test_frequency_sweep_stop_turns_rf_off_even_when_end_off_disabled(self) -> None:
        generator = FakeInstrument(query_responses={"*IDN?": "Rohde&Schwarz,SMIQ03B,123,1.0", ":SOUR:FREQ:CW?": "100000000", ":SOUR:POW?": "-30", ":OUTP?": "1"})
        measurement = FakeInstrument(query_responses={"*IDN?": "KEITHLEY INSTRUMENTS INC.,MODEL 2000,123,1.0"})
        calls = {"count": 0}

        def stop_after_generator_set() -> bool:
            calls["count"] += 1
            return calls["count"] >= 2

        result = run_frequency_sweep(
            generator,  # type: ignore[arg-type]
            measurement,  # type: ignore[arg-type]
            FrequencySweepConfig("100 MHz", "101 MHz", "1 MHz", "-30 dBm", 0, 0, "dmm", rf_off_at_end=False),
            stop_requested=stop_after_generator_set,
        )

        self.assertTrue(result.stopped)
        self.assertEqual(generator.writes[-1], ":OUTP OFF")

    def test_frequency_sweep_aborts_on_generator_error_and_turns_rf_off(self) -> None:
        generator = FakeInstrument(query_responses={"*IDN?": "Rohde&Schwarz,SMIQ03B,123,1.0", ":SOUR:FREQ:CW?": "100000000", ":SOUR:POW?": "-30", ":OUTP?": "1"})
        measurement = FakeInstrument(query_responses={"*IDN?": "KEITHLEY INSTRUMENTS INC.,MODEL 2000,123,1.0", ":READ?": "4.2"})
        original_write = generator.write

        def failing_write(command: str) -> None:
            if command == ":SOUR:FREQ:CW 100000000HZ":
                raise RuntimeError("generator failed")
            original_write(command)

        generator.write = failing_write  # type: ignore[method-assign]
        result = run_frequency_sweep(
            generator,  # type: ignore[arg-type]
            measurement,  # type: ignore[arg-type]
            FrequencySweepConfig("100 MHz", "101 MHz", "1 MHz", "-30 dBm", 0, 0, "dmm"),
        )

        self.assertEqual(result.actual_count, 1)
        self.assertEqual(result.error_count, 1)
        self.assertEqual(measurement.queries, ["*IDN?"])
        self.assertIn(":OUTP OFF", generator.writes)

    def test_frequency_sweep_export_charts_only_value(self) -> None:
        from openpyxl import load_workbook

        content = "SetFrequencyHz,Index,ElapsedSeconds,Value,Status\n100000000,1,0.1,4.2,OK\n101000000,2,0.2,4.4,OK\n"
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "results.xlsx"
            export = append_result(workbook_path, "GPIB0::1::INSTR", "Rohde&Schwarz,SMIQ03B,123,1.0", AcquisitionResult("frequency sweep", "csv", content))
            workbook = load_workbook(workbook_path)
            sheet = workbook[export.sheet_name]

            self.assertEqual(len(sheet._charts), 1)
            self.assertEqual(len(sheet._charts[0].series), 1)

    def test_voltage_sweep_sets_power_supply_and_reads_dmm(self) -> None:
        supply = FakeInstrument(
            query_responses={
                "*IDN?": "HAMEG,HMP4030,123,1.0",
                "VOLT?": "1.000",
                "CURR?": "0.100",
                "MEAS:VOLT?": "1.000",
                "MEAS:CURR?": "0.010",
                "OUTP:SEL?": "1",
                "OUTP:GEN?": "1",
            }
        )
        supply.address = "GPIB0::3::INSTR"
        measurement = FakeInstrument(query_responses={"*IDN?": "KEITHLEY INSTRUMENTS INC.,MODEL 2000,123,1.0", ":READ?": "0.5"})
        measurement.address = "GPIB0::2::INSTR"

        result = run_voltage_sweep(
            supply,  # type: ignore[arg-type]
            measurement,  # type: ignore[arg-type]
            VoltageSweepConfig(
                start_voltage="0 V",
                stop_voltage="1 V",
                step_voltage="1 V",
                current_limit="0.1 A",
                channel=1,
                max_voltage=5,
                max_current=1,
                settle_s=0,
                measurement_mode="dmm",
            ),
        )

        self.assertEqual(result.actual_count, 2)
        self.assertEqual(result.ok_count, 2)
        self.assertIn("SetVoltageV,Index", result.csv_content)
        self.assertIn("SupplyVoltageMeasured", result.csv_content)
        self.assertIn("SupplyCurrentMeasured", result.csv_content)
        self.assertIn("0.000000", result.csv_content)
        self.assertIn("1.000000", result.csv_content)
        self.assertEqual(supply.writes[-2:], ["OUTP:GEN 0", "INST:NSEL 1"])
        self.assertEqual(measurement.queries, ["*IDN?", ":READ?", ":READ?"])

    def test_timed_switch_generator_toggles_rf_and_ends_off(self) -> None:
        generator = FakeInstrument(query_responses={"*IDN?": "Rohde&Schwarz,SMIQ03B,123,1.0", ":SOUR:FREQ:CW?": "100000000", ":SOUR:POW?": "-30", ":OUTP?": "0"})
        generator.address = "GPIB0::1::INSTR"

        result = run_timed_switch(
            generator,  # type: ignore[arg-type]
            TimedSwitchConfig(source_type="generator", on_s=0.001, off_s=0, repetitions=2, end_off=True),
        )

        self.assertEqual(result.actual_count, 4)
        self.assertEqual(result.ok_count, 4)
        self.assertEqual(generator.writes, [":OUTP ON", ":OUTP OFF", ":OUTP ON", ":OUTP OFF", ":OUTP OFF"])
        self.assertIn("SourceType", result.csv_content)

    def test_timed_switch_power_supply_channel_mode(self) -> None:
        supply = FakeInstrument(
            query_responses={
                "*IDN?": "HAMEG,HMP4030,123,1.0",
                "VOLT?": "1.000",
                "CURR?": "0.100",
                "MEAS:VOLT?": "1.000",
                "MEAS:CURR?": "0.010",
                "OUTP:SEL?": "0",
                "OUTP:GEN?": "1",
            }
        )

        result = run_timed_switch(
            supply,  # type: ignore[arg-type]
            TimedSwitchConfig(source_type="power_supply", on_s=0.001, off_s=0, repetitions=1, power_supply_channel=2, power_supply_switch_mode="channel"),
        )

        self.assertEqual(result.actual_count, 2)
        self.assertEqual(supply.writes[:3], ["INST:NSEL 2", "INST:NSEL 2", "OUTP:SEL 1"])
        self.assertIn("OUTP:SEL 0", supply.writes)

    def test_timed_switch_rejects_generator_power_above_limit_before_writes(self) -> None:
        generator = FakeInstrument(query_responses={"*IDN?": "Rohde&Schwarz,SMIQ03B,123,1.0"})

        with self.assertRaises(ValueError):
            run_timed_switch(
                generator,  # type: ignore[arg-type]
                TimedSwitchConfig(source_type="generator", on_s=1, off_s=1, repetitions=1, generator_power="10 dBm", generator_max_power_dbm=0),
            )

        self.assertEqual(generator.writes, [])

    def test_timed_switch_rejects_current_generator_power_without_setup(self) -> None:
        generator = FakeInstrument(query_responses={"*IDN?": "Rohde&Schwarz,SMIQ03B,123,1.0", ":SOUR:FREQ:CW?": "100000000", ":SOUR:POW?": "10", ":OUTP?": "0"})

        with self.assertRaises(ValueError):
            run_timed_switch(
                generator,  # type: ignore[arg-type]
                TimedSwitchConfig(source_type="generator", on_s=1, off_s=1, repetitions=1, generator_power="-30 dBm", generator_max_power_dbm=0, setup_before_start=False),
            )

        self.assertEqual(generator.writes, [":OUTP OFF"])

    def test_frequency_sweep_records_final_off_failure(self) -> None:
        generator = FakeInstrument(query_responses={"*IDN?": "Rohde&Schwarz,SMIQ03B,123,1.0", ":SOUR:FREQ:CW?": "100000000", ":SOUR:POW?": "-30", ":OUTP?": "1"})
        measurement = FakeInstrument(query_responses={"*IDN?": "KEITHLEY INSTRUMENTS INC.,MODEL 2000,123,1.0"})

        def failing_write(command: str) -> None:
            if command == ":OUTP OFF":
                raise RuntimeError("off failed")
            generator.writes.append(command)

        generator.write = failing_write  # type: ignore[method-assign]
        result = run_frequency_sweep(
            generator,  # type: ignore[arg-type]
            measurement,  # type: ignore[arg-type]
            FrequencySweepConfig("100 MHz", "100 MHz", "1 MHz", "-30 dBm", 0, 0, "dmm"),
            stop_requested=lambda: True,
        )

        self.assertEqual(result.error_count, 1)
        self.assertIn("FINAL_OFF_ERROR", result.csv_content)

    def test_voltage_sweep_rejects_voltage_above_limit_before_writes(self) -> None:
        supply = FakeInstrument(query_responses={"*IDN?": "HAMEG,HMP4030,123,1.0"})
        measurement = FakeInstrument(query_responses={"*IDN?": "KEITHLEY INSTRUMENTS INC.,MODEL 2000,123,1.0"})

        with self.assertRaises(ValueError):
            run_voltage_sweep(
                supply,  # type: ignore[arg-type]
                measurement,  # type: ignore[arg-type]
                VoltageSweepConfig("0 V", "10 V", "1 V", "0.1 A", 1, 5, 1, 0, "dmm"),
            )

        self.assertEqual(supply.writes, [])

    def test_voltage_sweep_export_charts_only_value(self) -> None:
        from openpyxl import load_workbook

        content = "SetVoltageV,Index,ElapsedSeconds,Value,Status\n0,1,0.1,4.2,OK\n1,2,0.2,4.4,OK\n"
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "results.xlsx"
            export = append_result(workbook_path, "GPIB0::3::INSTR", "HAMEG,HMP4030,123,1.0", AcquisitionResult("voltage sweep", "csv", content))
            workbook = load_workbook(workbook_path)
            sheet = workbook[export.sheet_name]

            self.assertEqual(len(sheet._charts), 1)
            self.assertEqual(len(sheet._charts[0].series), 1)


if __name__ == "__main__":
    unittest.main()
