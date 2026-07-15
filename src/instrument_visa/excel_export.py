from __future__ import annotations

import csv
import threading
from dataclasses import dataclass
from datetime import datetime
from io import StringIO
from math import ceil, isfinite
from pathlib import Path
from uuid import uuid4
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .acquisition import AcquisitionResult


MAX_CHART_POINTS = 2000
CHART_WIDTH = 30
CHART_HEIGHT = 15
_WORKBOOK_LOCKS: dict[Path, threading.Lock] = {}
_WORKBOOK_LOCKS_LOCK = threading.Lock()


@dataclass(frozen=True)
class ExportResult:
    workbook_path: Path
    artifact_path: Path | None = None
    sheet_name: str | None = None


def append_result(workbook_path: Path, address: str, idn: str, result: AcquisitionResult, sheet_name: str | None = None) -> ExportResult:
    with _workbook_lock(workbook_path):
        return _append_result_unlocked(workbook_path, address, idn, result, sheet_name)


def _append_result_unlocked(workbook_path: Path, address: str, idn: str, result: AcquisitionResult, preferred_sheet_name: str | None = None) -> ExportResult:
    artifact: Path | None = None
    result_sheet_name: str | None = None
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    if workbook_path.exists():
        try:
            workbook = load_workbook(workbook_path)
        except (BadZipFile, KeyError):
            workbook_path = _backup_corrupt_workbook(workbook_path)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Results"
            _initialize_results_sheet(sheet)
        else:
            sheet = _get_or_create_results_sheet(workbook)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        _initialize_results_sheet(sheet)

    if isinstance(result.content, bytes):
        artifact = _artifact_path(workbook_path, result.file_type)
        artifact.write_bytes(result.content)
        image_sheet_name = _add_image_sheet(workbook, timestamp, address, idn, result, artifact)
        result_sheet_name = image_sheet_name
        target = f"sheet:{image_sheet_name}" if image_sheet_name else str(artifact)
        sheet.append([timestamp, address, idn, result.kind, result.file_type, target])
    elif result.file_type == "csv":
        if result.kind in {"34970A data logger", "34970A measurement plan"}:
            data_sheet = _get_or_create_sheet(workbook, preferred_sheet_name or "34970A Measurements")
        elif result.kind == "CA-410 measurement":
            data_sheet = _get_or_create_sheet(workbook, preferred_sheet_name or "CA-410 Measurements")
        else:
            data_sheet = workbook.create_sheet(_unique_sheet_name(workbook, result.kind))
        result_sheet_name = data_sheet.title
        sheet.append([timestamp, address, idn, result.kind, result.file_type, f"sheet:{result_sheet_name}"])
        if result.kind in {"34970A data logger", "34970A measurement plan"}:
            _append_34970a_csv(data_sheet, result.content)
        elif result.kind == "CA-410 measurement":
            _append_34970a_csv(data_sheet, result.content)
        else:
            data_sheet.append(["Timestamp", timestamp])
            data_sheet.append(["Address", address])
            data_sheet.append(["IDN", idn])
            data_sheet.append(["Kind", result.kind])
            data_sheet.append(["FileType", result.file_type])
            data_sheet.append([])
            _append_csv(data_sheet, result.content)
        if result.kind in {"waveform", "picoscope analog", "picoscope digital"}:
            _add_waveform_chart(data_sheet)
        elif result.kind in {"frequency sweep", "voltage sweep"}:
            _add_sweep_chart(data_sheet, result.kind)
        elif result.kind == "CA-410 measurement":
            _add_ca410_charts(data_sheet)
        _format_sheet(data_sheet)
    elif result.file_type.startswith("s") and result.file_type.endswith("p"):
        artifact = _artifact_path(workbook_path, result.file_type)
        artifact.write_text(result.content, encoding="utf-8")
        sheet.append([timestamp, address, idn, result.kind, result.file_type, str(artifact)])
    elif result.file_type == "txt":
        artifact = _artifact_path(workbook_path, result.file_type)
        artifact.write_text(str(result.content), encoding="utf-8")
        sheet.append([timestamp, address, idn, result.kind, result.file_type, str(artifact)])
    else:
        sheet.append([timestamp, address, idn, result.kind, result.file_type, _safe_excel_text(result.content)])

    _format_sheet(sheet)
    workbook.save(workbook_path)
    return ExportResult(workbook_path=workbook_path, artifact_path=artifact, sheet_name=result_sheet_name)


def _workbook_lock(workbook_path: Path) -> threading.Lock:
    resolved = workbook_path.expanduser().resolve(strict=False)
    with _WORKBOOK_LOCKS_LOCK:
        lock = _WORKBOOK_LOCKS.get(resolved)
        if lock is None:
            lock = threading.Lock()
            _WORKBOOK_LOCKS[resolved] = lock
        return lock


def _append_csv(sheet, content: str) -> None:
    for row in csv.reader(StringIO(content)):
        sheet.append([_coerce_excel_value(value) for value in row])


def _append_34970a_csv(sheet, content: str) -> None:
    rows = [[_coerce_excel_value(value) for value in row] for row in csv.reader(StringIO(content))]
    if not rows:
        return
    header = rows[0]
    if not any(cell.value is not None for row in sheet.iter_rows() for cell in row):
        for column, value in enumerate(header, start=1):
            sheet.cell(row=1, column=column).value = value
    existing_headers = [sheet.cell(row=1, column=column).value for column in range(1, sheet.max_column + 1)]
    for value in header:
        if value not in existing_headers:
            existing_headers.append(value)
            sheet.cell(row=1, column=len(existing_headers)).value = value
    columns_by_header = {value: index for index, value in enumerate(existing_headers, start=1)}
    for row in rows[1:]:
        output_row = [""] * len(existing_headers)
        for header_value, value in zip(header, row):
            output_row[columns_by_header[header_value] - 1] = value
        sheet.append(output_row)


def _get_or_create_sheet(workbook, name: str):
    if name in workbook.sheetnames:
        return workbook[name]
    return workbook.create_sheet(name[:31])


def _get_or_create_results_sheet(workbook):
    if "Results" in workbook.sheetnames:
        sheet = workbook["Results"]
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
            _initialize_results_sheet(sheet)
        return sheet
    sheet = workbook.create_sheet("Results", 0)
    _initialize_results_sheet(sheet)
    return sheet


def _coerce_excel_value(value: str) -> str | float:
    value = value.strip()
    if not value:
        return ""
    try:
        parsed = float(value)
    except ValueError:
        return _safe_excel_text(value)
    return parsed if isfinite(parsed) else _safe_excel_text(value)


def _safe_excel_text(value: object) -> str:
    text = ILLEGAL_CHARACTERS_RE.sub("", str(value))
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _initialize_results_sheet(sheet) -> None:
    sheet.append(["Timestamp", "Address", "IDN", "Kind", "FileType", "ValueOrFile"])


def _artifact_path(workbook_path: Path, file_type: str) -> Path:
    return workbook_path.with_name(f"{workbook_path.stem}-{uuid4().hex[:8]}.{file_type}")


def _add_image_sheet(workbook, timestamp: str, address: str, idn: str, result: AcquisitionResult, artifact: Path) -> str | None:
    if result.file_type.lower() not in {"png", "jpg", "jpeg"}:
        return None

    image_sheet = workbook.create_sheet(_unique_sheet_name(workbook, result.kind))
    image_sheet.append(["Timestamp", timestamp])
    image_sheet.append(["Address", address])
    image_sheet.append(["IDN", idn])
    image_sheet.append(["Kind", result.kind])
    image_sheet.append(["ImageFile", str(artifact)])
    image_sheet.append([])
    try:
        image = ExcelImage(str(artifact))
        image.anchor = "A7"
        image_sheet.add_image(image)
    except Exception as exc:
        image_sheet.append(["ImageEmbedError", str(exc)])
    _format_sheet(image_sheet)
    return image_sheet.title


def _add_waveform_chart(sheet) -> None:
    header_row = _find_waveform_header_row(sheet)
    if header_row is None:
        return

    scan_max_row = min(sheet.max_row, header_row + MAX_CHART_POINTS)
    numeric_columns = _numeric_columns(sheet, header_row, scan_max_row)
    if not numeric_columns:
        return

    header_values = [sheet.cell(header_row, column).value for column in range(1, sheet.max_column + 1)]
    first_header = str(header_values[0]).strip().lower() if header_values and header_values[0] is not None else ""
    has_x_axis = 1 in numeric_columns and first_header in {"time", "frequency", "freq", "setfrequencyhz", "[hz]", "hz"}
    max_row = _last_numeric_row(sheet, header_row + 1, numeric_columns)
    if max_row <= header_row:
        return

    header_row, max_row, numeric_columns, anchor = _add_chart_data_table(sheet, header_row, max_row, numeric_columns, has_x_axis)
    _force_excel_recalculation(sheet)

    if has_x_axis and len(numeric_columns) > 1:
        _add_scatter_waveform_chart(sheet, header_row, max_row, numeric_columns, anchor)
    else:
        _add_line_waveform_chart(sheet, header_row, max_row, numeric_columns, anchor)


def _add_sweep_chart(sheet, kind: str) -> None:
    x_header = "setfrequencyhz" if kind == "frequency sweep" else "setvoltagev"
    header_row = _find_sweep_header_row(sheet, x_header)
    if header_row is None:
        return
    headers = {str(sheet.cell(header_row, column).value or "").strip().lower(): column for column in range(1, sheet.max_column + 1)}
    x_column = headers.get(x_header)
    y_column = headers.get("value")
    if x_column is None or y_column is None:
        return
    max_row = _last_numeric_row(sheet, header_row + 1, [y_column])
    if max_row <= header_row:
        return
    if not any(isinstance(sheet.cell(row, y_column).value, (int, float)) for row in range(header_row + 1, max_row + 1)):
        return

    x_column, y_column, header_row, max_row = _sweep_chart_columns(sheet, header_row, max_row, x_column, y_column, x_header)

    chart = ScatterChart()
    chart.title = "Frequency Sweep" if kind == "frequency sweep" else "Voltage Sweep"
    chart.x_axis.title = "Frequency [Hz]" if kind == "frequency sweep" else "Voltage [V]"
    chart.y_axis.title = "Value"
    chart.legend = None
    chart.width = CHART_WIDTH
    chart.height = CHART_HEIGHT
    x_values = Reference(sheet, min_col=x_column, min_row=header_row + 1, max_row=max_row)
    y_values = Reference(sheet, min_col=y_column, min_row=header_row + 1, max_row=max_row)
    chart.series.append(Series(y_values, x_values, title="Value"))
    sheet.add_chart(chart, _chart_anchor(sheet))


def _add_ca410_charts(sheet) -> None:
    header_row = _find_ca410_header_row(sheet)
    if header_row is None:
        return
    headers = {str(sheet.cell(header_row, column).value or "").strip().lower(): column for column in range(1, sheet.max_column + 1)}
    first_data_row = header_row + 1
    value_columns = [column for name, column in headers.items() if name in {"x", "y", "lv", "tcp", "duv", "fmaflickerpercent", "tempshift"}]
    max_row = _last_numeric_row(sheet, first_data_row, value_columns)
    if max_row <= header_row:
        return

    sheet._charts = []
    anchor_column = get_column_letter(min(sheet.max_column + 2, 14))
    _add_ca410_white_point_chart(sheet, headers, header_row, max_row, f"{anchor_column}2")
    _add_ca410_line_chart(sheet, headers, header_row, max_row, ["lv"], "CA-410 Leuchtdichte", "Lv", f"{anchor_column}20")
    _add_ca410_line_chart(sheet, headers, header_row, max_row, ["fmaflickerpercent", "tempshift"], "CA-410 Flicker / TempShift", "Value", f"{anchor_column}38")
    _add_ca410_line_chart(sheet, headers, header_row, max_row, ["tcp", "duv"], "CA-410 Farbtemperatur / duv", "Value", f"{anchor_column}56")


def _add_ca410_white_point_chart(sheet, headers: dict[str, int], header_row: int, max_row: int, anchor: str) -> None:
    x_column = headers.get("x")
    y_column = headers.get("y")
    if x_column is None or y_column is None:
        return
    if not any(isinstance(sheet.cell(row, x_column).value, (int, float)) and isinstance(sheet.cell(row, y_column).value, (int, float)) for row in range(header_row + 1, max_row + 1)):
        return
    chart = ScatterChart()
    chart.title = "CA-410 Weißpunkt / Farbort"
    chart.x_axis.title = "x"
    chart.y_axis.title = "y"
    chart.legend = None
    chart.width = 18
    chart.height = 14
    try:
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = 1
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1
    except AttributeError:
        pass
    x_values = Reference(sheet, min_col=x_column, min_row=header_row + 1, max_row=max_row)
    y_values = Reference(sheet, min_col=y_column, min_row=header_row + 1, max_row=max_row)
    chart.series.append(Series(y_values, x_values, title="x/y"))
    sheet.add_chart(chart, anchor)


def _add_ca410_line_chart(sheet, headers: dict[str, int], header_row: int, max_row: int, header_names: list[str], title: str, y_axis_title: str, anchor: str) -> None:
    columns = [headers[name] for name in header_names if name in headers and any(isinstance(sheet.cell(row, headers[name]).value, (int, float)) for row in range(header_row + 1, max_row + 1))]
    if not columns:
        return
    chart = LineChart()
    chart.title = title
    chart.x_axis.title = "Messung"
    chart.y_axis.title = y_axis_title
    chart.legend.position = "r"
    chart.width = 24
    chart.height = 12
    for column in columns:
        data = Reference(sheet, min_col=column, min_row=header_row, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
    timestamp_column = headers.get("timestamp")
    if timestamp_column is not None:
        categories = Reference(sheet, min_col=timestamp_column, min_row=header_row + 1, max_row=max_row)
        chart.set_categories(categories)
    sheet.add_chart(chart, anchor)


def _sweep_chart_columns(sheet, header_row: int, max_row: int, x_column: int, y_column: int, x_header: str) -> tuple[int, int, int, int]:
    source_rows = [row for row in range(header_row + 1, max_row + 1) if isinstance(sheet.cell(row, y_column).value, (int, float))]
    if len(source_rows) <= MAX_CHART_POINTS:
        return x_column, y_column, header_row, max_row

    rows_per_bucket = max(1, ceil(len(source_rows) / MAX_CHART_POINTS))
    sampled_rows = source_rows[::rows_per_bucket]
    if sampled_rows[-1] != source_rows[-1]:
        sampled_rows.append(source_rows[-1])

    target_x_column = sheet.max_column + 2
    target_y_column = target_x_column + 1
    target_header_row = header_row
    sheet.cell(target_header_row, target_x_column).value = "SetFrequencyHz" if x_header == "setfrequencyhz" else "SetVoltageV"
    sheet.cell(target_header_row, target_y_column).value = "Value"
    for cell in (sheet.cell(target_header_row, target_x_column), sheet.cell(target_header_row, target_y_column)):
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAD3")

    note = sheet.cell(max(1, header_row - 1), target_x_column)
    note.value = f"Diagramm-Extrakt: jeder {rows_per_bucket}. numerische Messpunkt plus letzter Punkt; vollständige Daten links."
    note.font = Font(bold=True)
    note.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    note.alignment = Alignment(wrap_text=True)

    target_row = target_header_row + 1
    for source_row in sampled_rows:
        sheet.cell(target_row, target_x_column).value = f"={get_column_letter(x_column)}{source_row}"
        sheet.cell(target_row, target_y_column).value = f"={get_column_letter(y_column)}{source_row}"
        target_row += 1
    _force_excel_recalculation(sheet)
    return target_x_column, target_y_column, target_header_row, target_row - 1


def _find_waveform_header_row(sheet) -> int | None:
    for row in range(1, sheet.max_row + 1):
        values = [sheet.cell(row, column).value for column in range(1, sheet.max_column + 1)]
        normalized = {str(value).strip().lower() for value in values if value is not None}
        if normalized & {"time", "frequency", "trace", "[hz]", "hz", "ch1", "channel1"}:
            return row
    return None


def _find_sweep_header_row(sheet, x_header: str) -> int | None:
    for row in range(1, sheet.max_row + 1):
        values = [sheet.cell(row, column).value for column in range(1, sheet.max_column + 1)]
        normalized = {str(value).strip().lower() for value in values if value is not None}
        if {x_header, "value"}.issubset(normalized):
            return row
    return None


def _find_ca410_header_row(sheet) -> int | None:
    for row in range(1, sheet.max_row + 1):
        values = [sheet.cell(row, column).value for column in range(1, sheet.max_column + 1)]
        normalized = {str(value).strip().lower() for value in values if value is not None}
        if {"timestamp", "status", "probe", "displaymode"}.issubset(normalized) and ({"x", "y", "lv"}.issubset(normalized) or "tcp" in normalized):
            return row
    return None


def _last_numeric_row(sheet, first_data_row: int, numeric_columns: list[int]) -> int:
    for row in range(sheet.max_row, first_data_row - 1, -1):
        if any(isinstance(sheet.cell(row, column).value, (int, float)) for column in numeric_columns):
            return row
    return first_data_row - 1


def _numeric_columns(sheet, header_row: int, max_row: int) -> list[int]:
    columns: list[int] = []
    for column in range(1, sheet.max_column + 1):
        header = str(sheet.cell(header_row, column).value or "").strip().lower()
        if header in {"index"}:
            continue
        if any(isinstance(sheet.cell(row, column).value, (int, float)) for row in range(header_row + 1, max_row + 1)):
            columns.append(column)
    return columns


def _add_chart_data_table(sheet, source_header_row: int, source_max_row: int, source_numeric_columns: list[int], has_x_axis: bool) -> tuple[int, int, list[int], str]:
    source_columns = [1, *[column for column in source_numeric_columns if column != 1]] if has_x_axis else source_numeric_columns
    value_columns = [column for column in source_numeric_columns if not (has_x_axis and column == 1)]
    start_column = sheet.max_column + 2
    toggle_row = max(1, source_header_row - 2)
    note_row = max(1, source_header_row - 1)
    sample_header_row = source_header_row if source_header_row > 1 else 2
    sample_data_row = sample_header_row + 1
    source_data_rows = source_max_row - source_header_row
    is_sampled = source_data_rows > MAX_CHART_POINTS
    rows_per_bucket = _chart_bucket_size(source_data_rows, len(value_columns)) if is_sampled else 1

    _add_channel_toggles(sheet, toggle_row, start_column, source_columns, has_x_axis)

    note = sheet.cell(note_row, start_column)
    note.value = (
        f"Diagramm-Extrakt: Min/Max je {rows_per_bucket} Punkte plus letzter Punkt; vollständige Daten links."
        if is_sampled
        else "Diagramm-Daten: vollständiger Datenbereich; Kanäle über TRUE/FALSE-Schalter ein-/ausblendbar."
    )
    note.font = Font(bold=True)
    note.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    note.alignment = Alignment(wrap_text=True)

    for offset, source_column in enumerate(source_columns):
        cell = sheet.cell(sample_header_row, start_column + offset)
        cell.value = sheet.cell(source_header_row, source_column).value or f"Column {source_column}"
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAD3")

    copied_rows: set[int] = set()
    for source_row in _chart_source_rows(sheet, source_header_row, source_max_row, rows_per_bucket, value_columns or source_numeric_columns, is_sampled):
        if source_row in copied_rows:
            continue
        _copy_chart_row(sheet, source_row, sample_data_row, source_columns, start_column, toggle_row, has_x_axis)
        copied_rows.add(source_row)
        sample_data_row += 1

    if source_max_row not in copied_rows:
        _copy_chart_row(sheet, source_max_row, sample_data_row, source_columns, start_column, toggle_row, has_x_axis)
        sample_data_row += 1

    sample_columns = list(range(start_column, start_column + len(source_columns)))
    anchor = f"{get_column_letter(start_column + len(source_columns) + 2)}{source_header_row}"
    return sample_header_row, sample_data_row - 1, sample_columns, anchor


def _add_channel_toggles(sheet, toggle_row: int, start_column: int, source_columns: list[int], has_x_axis: bool) -> None:
    label = sheet.cell(toggle_row, start_column)
    label.value = "Diagramm-Kanäle"
    label.font = Font(bold=True)
    label.fill = PatternFill(fill_type="solid", fgColor="D9EAD3")

    validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    sheet.add_data_validation(validation)
    for offset, source_column in enumerate(source_columns):
        target_column = start_column + offset
        toggle = sheet.cell(toggle_row, target_column)
        if has_x_axis and offset == 0:
            toggle.value = "X-Achse"
        else:
            toggle.value = True
            validation.add(toggle)
        toggle.font = Font(bold=True)
        toggle.fill = PatternFill(fill_type="solid", fgColor="D9EAD3")


def _chart_source_rows(sheet, source_header_row: int, source_max_row: int, rows_per_bucket: int, value_columns: list[int], is_sampled: bool) -> list[int]:
    if not is_sampled:
        return [
            row
            for row in range(source_header_row + 1, source_max_row + 1)
            if any(isinstance(sheet.cell(row, column).value, (int, float)) for column in value_columns)
        ]

    rows: list[int] = []
    for bucket_start in range(source_header_row + 1, source_max_row + 1, rows_per_bucket):
        bucket_end = min(bucket_start + rows_per_bucket - 1, source_max_row)
        rows.extend(_bucket_extreme_rows(sheet, bucket_start, bucket_end, value_columns))
    return rows


def _copy_chart_row(sheet, source_row: int, sample_row: int, source_columns: list[int], start_column: int, toggle_row: int, has_x_axis: bool) -> None:
    for offset, source_column in enumerate(source_columns):
        target_column = start_column + offset
        source_ref = f"{get_column_letter(source_column)}{source_row}"
        if has_x_axis and offset == 0:
            sheet.cell(sample_row, target_column).value = f"={source_ref}"
        else:
            toggle_ref = f"{get_column_letter(target_column)}${toggle_row}"
            sheet.cell(sample_row, target_column).value = f"=IF({toggle_ref},{source_ref},NA())"


def _chart_bucket_size(source_data_rows: int, value_column_count: int) -> int:
    rows_per_bucket_output = max(2, value_column_count * 2 + 2)
    bucket_count = max(1, MAX_CHART_POINTS // rows_per_bucket_output)
    return max(1, ceil(source_data_rows / bucket_count))


def _bucket_extreme_rows(sheet, start_row: int, end_row: int, value_columns: list[int]) -> list[int]:
    rows: set[int] = {start_row, end_row}
    for column in value_columns:
        numeric_values = [
            (row, sheet.cell(row, column).value)
            for row in range(start_row, end_row + 1)
            if isinstance(sheet.cell(row, column).value, (int, float))
        ]
        if not numeric_values:
            continue
        rows.add(min(numeric_values, key=lambda item: item[1])[0])
        rows.add(max(numeric_values, key=lambda item: item[1])[0])
    return sorted(rows)


def _add_scatter_waveform_chart(sheet, header_row: int, max_row: int, numeric_columns: list[int], anchor: str) -> None:
    chart = ScatterChart()
    chart.title = "Waveform"
    x_column = numeric_columns[0]
    chart.x_axis.title = _axis_title(str(sheet.cell(header_row, x_column).value or "X"), axis="x")
    chart.y_axis.title = _waveform_y_axis_title(sheet, header_row, numeric_columns[1:])
    chart.legend.position = "r"
    chart.width = CHART_WIDTH
    chart.height = CHART_HEIGHT

    x_values = Reference(sheet, min_col=x_column, min_row=header_row + 1, max_row=max_row)
    for column in numeric_columns:
        if column == x_column:
            continue
        y_values = Reference(sheet, min_col=column, min_row=header_row + 1, max_row=max_row)
        series = Series(y_values, x_values, title=str(sheet.cell(header_row, column).value or f"Series {column}"))
        chart.series.append(series)

    if chart.series:
        sheet.add_chart(chart, anchor)


def _add_line_waveform_chart(sheet, header_row: int, max_row: int, numeric_columns: list[int], anchor: str) -> None:
    chart = LineChart()
    chart.title = "Waveform"
    chart.x_axis.title = "Point"
    chart.y_axis.title = _waveform_y_axis_title(sheet, header_row, numeric_columns)
    chart.legend.position = "r"
    chart.width = CHART_WIDTH
    chart.height = CHART_HEIGHT

    data = Reference(sheet, min_col=min(numeric_columns), max_col=max(numeric_columns), min_row=header_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    sheet.add_chart(chart, anchor)


def _chart_anchor(sheet) -> str:
    return f"{get_column_letter(min(sheet.max_column + 2, 12))}7"


def _axis_title(header: str, axis: str) -> str:
    normalized = header.strip().lower()
    if axis == "x" and normalized == "time":
        return "Time [s]"
    if axis == "x" and normalized in {"frequency", "freq", "setfrequencyhz", "[hz]", "hz"}:
        return "Frequency [Hz]"
    return header or axis.upper()


def _waveform_y_axis_title(sheet, header_row: int, value_columns: list[int]) -> str:
    headers = [str(sheet.cell(header_row, column).value or "").strip().lower() for column in value_columns]
    if any("dbm" in header for header in headers):
        return "Level [dBm]"
    if any(header.startswith("ch") for header in headers):
        return "Amplitude"
    return "Value"


def _force_excel_recalculation(sheet) -> None:
    try:
        sheet.parent.calculation.fullCalcOnLoad = True
        sheet.parent.calculation.forceFullCalc = True
        sheet.parent.calculation.calcMode = "auto"
    except AttributeError:
        pass


def _unique_sheet_name(workbook, kind: str) -> str:
    prefix = "".join(character for character in kind.title() if character.isalnum())[:16] or "Result"
    timestamp = datetime.now().strftime("%H%M%S")
    base_name = f"{prefix}-{timestamp}"[:31]
    sheet_name = base_name
    counter = 1
    while sheet_name in workbook.sheetnames:
        suffix = f"-{counter}"
        sheet_name = f"{base_name[:31 - len(suffix)]}{suffix}"
        counter += 1
    return sheet_name


def _backup_corrupt_workbook(workbook_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = workbook_path.with_name(f"{workbook_path.stem}.corrupt-{timestamp}{workbook_path.suffix}")
    workbook_path.replace(backup_path)
    return workbook_path


def _format_sheet(sheet) -> None:
    waveform_header_row = _find_waveform_header_row(sheet)
    sheet.freeze_panes = f"A{waveform_header_row + 1}" if waveform_header_row else "A2"
    if waveform_header_row:
        original_max_column = _original_data_max_column(sheet, waveform_header_row)
        if original_max_column > 0:
            sheet.auto_filter.ref = f"A{waveform_header_row}:{get_column_letter(original_max_column)}{sheet.max_row}"
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells[:200]:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)


def _original_data_max_column(sheet, header_row: int) -> int:
    max_column = 0
    for column in range(1, sheet.max_column + 1):
        value = sheet.cell(header_row, column).value
        if value is None or str(value).strip() == "":
            break
        max_column = column
    return max_column
